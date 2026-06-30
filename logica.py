"""
Lógica do cruzamento Razão × SPED (Triple Check + CNAE + CST + Pendências).
Port da implementação client-side JS para Python — mantém todos os comportamentos:

  • Extração de NF do Complemento Histórico (com fallback pra coluna 15)
  • Mapa Nome→CNPJ derivado dos SPEDs
  • Enriquecimento Razão Social oficial + 1º/2º CNAE via BrasilAPI/CNPJa
  • Filtro CST 50–67 em A100 e F100 (faixas "Direito a Crédito" e "Crédito Presumido")
  • Triple Check: NF + CNPJ + Período (C100-EFD/A100), CNPJ + Período + Valor (F100)
  • Tolerância de R$ 100 para considerar valor batido
  • Coluna ANÁLISE DO CRUZAMENTO (texto enxuto: bloco encontrado ou valor da diferença)
  • Aba PENDÊNCIAS DE CRUZAMENTO (consolidação por CNPJ, ordenada DESC)
  • Padrão visual EFCT: Exo 2 + paleta navy/lime + sem gridlines
"""
import io
import re
import time
import difflib
import logging
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


log = logging.getLogger(__name__)

# ============================================================
# CONSTANTES
# ============================================================
TOLERANCIA_VALOR = 100.00
CNAE_CONCURRENCY = 3  # paralelismo conservador pra não bater rate limit
CNAE_TIMEOUT = 12  # segundos por request HTTP
USER_AGENT = 'Mozilla/5.0 EFCT-Hub/1.0'

# Cache global em memória (vive enquanto o processo Flask estiver de pé)
CNAE_CACHE: dict[str, dict] = {}

# Cores EFCT (paleta oficial)
COR_NAVY = '0C2B38'
COR_LIME = 'B3BC2B'
COR_AMBAR = '92400E'
COR_VERMELHO = '991B1B'
COR_BRANCO = 'FFFFFF'
FONT_BASE = 'Exo 2'

# ============================================================
# ESTILOS EFCT (openpyxl)
# ============================================================
def make_styles():
    """Cria os objetos de estilo openpyxl. Chamado uma vez por execução."""
    return {
        # Cabeçalho cols originais (1-15): fundo navy + texto lime
        'header': {
            'font': Font(name=FONT_BASE, size=12, color=COR_LIME),
            'fill': PatternFill('solid', fgColor=COR_NAVY),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        },
        # Cabeçalho cols adicionadas (16+): fundo lime + texto navy
        'new_header': {
            'font': Font(name=FONT_BASE, size=12, color=COR_NAVY),
            'fill': PatternFill('solid', fgColor=COR_LIME),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        },
        # Info text (NF, Razão Social, CNPJ, CNAEs) — branco + navy
        'info': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        },
        # Valor OK — branco + navy (alinhado à direita, formato número)
        'value_ok': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '#,##0.00',
        },
        # Valor divergente — branco + âmbar bold
        'value_warn': {
            'font': Font(name=FONT_BASE, size=11, color=COR_AMBAR, bold=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '#,##0.00',
        },
        # Alerta (ambíguo / não identificado) — branco + vermelho italic
        'alert': {
            'font': Font(name=FONT_BASE, size=11, color=COR_VERMELHO, italic=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        },
        # Análise OK
        'analise_ok': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        },
        # Análise divergente — só o valor
        'analise_warn': {
            'font': Font(name=FONT_BASE, size=11, color=COR_AMBAR, bold=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='right', vertical='center'),
        },
        # Análise não localizada
        'analise_err': {
            'font': Font(name=FONT_BASE, size=11, color=COR_VERMELHO, italic=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        },
        # Padrão das células de dados (todas em Exo 2 navy)
        'default': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(vertical='center'),
        },
        # Valor sem crédito (NF presente mas em CST fora de 50-67) — italic gray-navy
        'value_info': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY, italic=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '#,##0.00',
        },
        # Análise sem crédito — italic navy
        'analise_info': {
            'font': Font(name=FONT_BASE, size=11, color=COR_NAVY, italic=True),
            'fill': PatternFill('solid', fgColor=COR_BRANCO),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        },
    }


def apply_style(cell, style):
    """Aplica um dicionário de atributos a uma célula openpyxl."""
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'number_format' in style:
        cell.number_format = style['number_format']


# ============================================================
# UTILITÁRIOS DE NORMALIZAÇÃO
# ============================================================
def normalize_nf(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ''
    s = str(v).strip()
    if not s:
        return ''
    # Limpa decimais herdados do pandas (ex: "40273.0")
    if re.match(r'^\d+\.0+$', s):
        s = s.split('.')[0]
    if s.isdigit():
        return str(int(s))
    return s


def normalize_name(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ''
    nfd = unicodedata.normalize('NFD', str(s))
    no_accent = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', no_accent.upper())


# Sufixos societários e palavras genéricas que NÃO ajudam a identificar a empresa.
# São ignorados ao comparar nomes (senão "FULANO LTDA" e "BELTRANO LTDA" pareceriam
# similares só por causa do "LTDA").
_NAME_STOPWORDS = {
    'LTDA', 'ME', 'EPP', 'EIRELI', 'SA', 'S', 'A', 'EI', 'MEI', 'CIA',
    'E', 'DE', 'DA', 'DO', 'DOS', 'DAS', 'EM', 'COMERCIAL', 'COMERCIO',
    'INDUSTRIA', 'SERVICOS', 'SERVICO', 'LTD',
}


def _name_tokens(s) -> set:
    """Tokens significativos de um nome (sem acento, uppercase, sem stopwords)."""
    norm = _normalize_for_keyword(str(s)) if s else ''
    toks = re.findall(r'[A-Z0-9]+', norm)
    return {t for t in toks if t not in _NAME_STOPWORDS and len(t) > 1}


def _names_reconcile(a, b) -> float:
    """Mede o quanto dois nomes de empresa 'fecham' (0.0 a 1.0).

    Combina 3 sinais e pega o maior:
      • igualdade exata após normalização → 1.0
      • um nome contido no outro (substring longa) → 0.92
      • Jaccard dos tokens significativos (ignora LTDA/ME/etc)
      • similaridade Levenshtein da string compactada

    Usado pra GARANTIR que o CNPJ/razão social trazido fecha com o que está no
    Complemento Histórico — se não fechar, o CNPJ é descartado.
    """
    na, nb = normalize_name(a), normalize_name(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if len(na) >= 6 and (na in nb or nb in na):
        return 0.92
    ta, tb = _name_tokens(a), _name_tokens(b)
    jac = len(ta & tb) / len(ta | tb) if (ta and tb) else 0.0
    seq = difflib.SequenceMatcher(None, na, nb).ratio()
    return max(jac, seq)


# Limiar mínimo de reconciliação pra aceitar um CNPJ de fonte "fraca"
# (fuzzy/prefixo/web). Abaixo disso, o CNPJ é descartado (fica em branco).
RECONCILE_THRESHOLD = 0.60
# Fontes confiáveis por construção (o nome/CNPJ já está literalmente no histórico
# ou bate exatamente com o mapa SPED) — dispensam o portão de validação.
_TRUSTED_VIAS = {'DIRETO_HISTORICO', 'HISTORICO_SCAN', 'EXACT', 'EXACT_STRIPPED', 'STRIP_DIGITS'}


def normalize_cnpj(c) -> str:
    if c is None or (isinstance(c, float) and pd.isna(c)):
        return ''
    digits = re.sub(r'\D', '', str(c))
    if not digits:
        return ''
    return digits.zfill(14)


def format_cnpj(c) -> str:
    d = normalize_cnpj(c)
    if len(d) != 14:
        return d
    return f'{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}'


def to_month_year(val) -> str:
    """Normaliza qualquer formato de período para 'MM/YYYY'."""
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return ''
    if isinstance(val, (datetime, pd.Timestamp)):
        return f'{val.month:02d}/{val.year}'
    s = str(val).strip()
    # DD/MM/YYYY
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f'{m.group(2).zfill(2)}/{m.group(3)}'
    # YYYY-MM-DD
    m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f'{m.group(2).zfill(2)}/{m.group(1)}'
    # Excel serial (número)
    try:
        n = float(s.replace(',', '.'))
        d = datetime(1899, 12, 30) + timedelta(days=n)
        return f'{d.month:02d}/{d.year}'
    except (ValueError, OverflowError):
        pass
    return s


def to_number(v) -> float:
    if v is None or v == '' or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # Trata números BR ("1.234,56") e EN ("1234.56")
    if ',' in s and '.' in s:
        # 1.234,56
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    s = re.sub(r'[^\d\-\.]', '', s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def _most_common_cnpj(values) -> str | None:
    """Devolve o CNPJ válido (14 dígitos) mais frequente numa lista (ou None)."""
    counts = {}
    for v in values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        c = normalize_cnpj(v)
        if len(c) == 14:
            counts[c] = counts.get(c, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda x: x[1])[0]


def is_credit_cst(cst) -> bool:
    """Faixa CST 50–67 (Direito a Crédito + Crédito Presumido)."""
    if cst is None or cst == '' or (isinstance(cst, float) and pd.isna(cst)):
        return False
    s = str(cst).strip()
    if not s:
        return False
    try:
        # Pode vir como "50", "50.0", "50 - ...", etc.
        n = int(float(re.match(r'\s*(\d+)', s).group(1)))
    except (ValueError, AttributeError):
        return False
    return 50 <= n <= 67


# ============================================================
# EXTRAÇÃO DE HISTÓRICO
# ============================================================
# Padrões de extração de NF do Complemento Histórico — testados em ordem de confiança.
# Cada padrão captura o número que vem após uma palavra-chave fiscal.
NF_PATTERNS = [
    # 1. VlrrefNF<digits> — padrão específico de alguns sistemas contábeis
    re.compile(r'Vlr\s*ref\s*N[\.\s]*F[\.\s]*(\d{2,10})', re.IGNORECASE),
    # 2. N(F) / Nota (F)iscal / NFe / NF-e — aceita separadores ENTRE N e F também
    #    (cobre "N F 4480", "N.F 4480", "NF 4480", "Nota Fiscal 4480", "NFe 4480")
    re.compile(
        r'\bN(?:ota)?[\s\.\-]*F(?:iscal)?[\s\.\-/#:]*(?:e[\s\.\-/]*)?(?:n[ºo°]?\.?[\s]*)?(\d{2,10})\b',
        re.IGNORECASE,
    ),
    # 3. N° / N. / Nº (sem o F) — quando o histórico abrevia
    re.compile(r'\bN(?:[oOºo°]|\.)\s*[\s\.\-/#:]*(\d{2,10})\b', re.IGNORECASE),
    # 4. Doc / Documento / Doc Fiscal
    re.compile(
        r'\bDoc(?:umento)?(?:\s*fiscal)?[\s\.\-/#:]*(?:n[ºo°]?\.?\s*)?(\d{2,10})\b',
        re.IGNORECASE,
    ),
    # 5. Fatura / Duplicata
    re.compile(
        r'\b(?:Fatura|Duplicata|Dupl?)[\s\.\-/#:]*(?:n[ºo°]?\.?\s*)?(\d{2,10})\b',
        re.IGNORECASE,
    ),
    # 6. Boleto / Bol — comum quando o histórico é "Pagto boleto NF 12345"
    re.compile(r'\bBol(?:eto)?[\s\.\-/#:]*(?:n[ºo°]?\.?\s*)?(\d{2,10})\b', re.IGNORECASE),
    # 7. Número solto no INÍCIO do histórico seguido de nome (ex: "32029  ALLFOR SOLUCOES")
    #    Vários históricos não têm prefixo "NF" — só o número e o nome do fornecedor.
    re.compile(r'^\s*(\d{2,10})\s+[A-Za-zÀ-ÿ]'),
]

# CNPJ formatado (12.345.678/0001-90) ou 14 dígitos seguidos
RE_CNPJ_FORMATADO = re.compile(r'\b(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\b')
RE_CNPJ_RAW = re.compile(r'(?<!\d)(\d{14})(?!\d)')


def _looks_like_nf_number(digits: str) -> bool:
    """NF típica tem 2-10 dígitos. Acima disso geralmente é CPF/CNPJ/CEP/conta."""
    return bool(digits) and digits.isdigit() and 2 <= len(digits) <= 10


# Palavras de contexto fiscal que NUNCA fazem parte do nome do fornecedor.
# Comparadas após uppercase + remoção de acentos.
FISCAL_KEYWORDS = {
    'VLR', 'VALOR', 'PGTO', 'PAGTO', 'PAGAMENTO', 'PGMTO', 'PAGO',
    'REF', 'REFER', 'REFERENTE', 'REFERENCIA', 'REFNF',
    'NF', 'NFE', 'NOTA', 'FISCAL',
    'NFSE', 'NFS', 'NFSES', 'NFCE', 'NFC', 'DANFE', 'RPS', 'CTE', 'CT',  # tipos de documento fiscal
    'NO', 'NUM', 'NUMERO', 'NR', 'NRO', 'NUMERACAO', 'SERIE',  # abreviações de "Número" (No, Num, Nº já cobertos pelo upper_norm)
    'DOC', 'DOCUMENTO', 'DOCTO',
    'FATURA', 'DUPLICATA', 'DUPL',
    'BOLETO', 'BOL',
    'COMPRA', 'AQUISICAO', 'DEVOLUCAO', 'PRESTACAO',
    'VLRREF', 'VLRREFNF',
    'TOTAL', 'PARCELA', 'PARC',
    'CONFORME', 'CONF', 'CFE', 'CF',  # "conforme" abreviado
    'DESPESA', 'DESPESAS', 'LANCAMENTO', 'AVISO', 'AVS',  # prefixos de descrição
    'LR',  # "Lançamento Razão" — prefixo comum
    'APROP', 'APROPRIACAO',  # "Apropriação"
    'ORIGINAL', 'OMISSO', 'OMISSA',  # marcadores de status do lançamento
    'MENSALIDADE', 'MENSAL', 'AREA',  # descrições recorrentes (não fazem parte do nome)
    'RS',  # "R$" sem o cifrão (ex: "RS 1.200,00")
}
# Nota: 'SERVICO'/'SERVICOS' NÃO entram aqui — aparecem muito em razões sociais
# ("CLINISEG SERVICOS DE APOIO"). O prefixo "PRESTACAO DE SERVICOS" é tratado
# via stripping de conectores nas pontas do nome (ver smart_extract).

# Letras isoladas que costumam ser marcadores ("N F 12345"), não nomes
NF_MARKERS_SINGLE = {'N', 'F', 'NF', 'NFE'}

# Preposições/conectores que NUNCA iniciam uma razão social. São removidos só
# das PONTAS do nome extraído (no meio permanecem: "SERVICOS DE APOIO").
NAME_EDGE_STOPWORDS = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E', 'EM', 'A', 'O'}

# Palavras que indicam o INÍCIO de uma razão social por tipo de entidade.
# Usadas na estratégia "nome a partir do tipo de empresa" (ex: a razão começa em
# "FUNDO DE INVESTIMENTO...", "BANCO...", "CONDOMINIO...", "ASSOCIACAO...").
ANCORA_EMPRESA = {
    'FUNDO', 'BANCO', 'COMPANHIA', 'CIA', 'COOPERATIVA', 'COOP', 'ASSOCIACAO',
    'INSTITUTO', 'FUNDACAO', 'CONDOMINIO', 'SINDICATO', 'IGREJA', 'PAROQUIA',
    'MUNICIPIO', 'PREFEITURA', 'ESTADO', 'CARTORIO', 'CAIXA', 'CLINICA',
    'HOSPITAL', 'ESCOLA', 'COLEGIO', 'UNIVERSIDADE', 'FACULDADE', 'CENTRO',
    'SOCIEDADE', 'EMPRESA', 'AUTO', 'POSTO', 'SUPERMERCADO', 'MERCADO',
    'COMERCIAL', 'INDUSTRIA', 'INDUSTRIAL', 'TRANSPORTES', 'TRANSPORTADORA',
    'CONSTRUTORA', 'IMOBILIARIA', 'DISTRIBUIDORA', 'ATACADO', 'ATACADAO',
}


def _normalize_for_keyword(s: str) -> str:
    """Uppercase + remove acentos pra comparar com FISCAL_KEYWORDS."""
    nfd = unicodedata.normalize('NFD', s.upper())
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')


def _preprocess_historico(s: str) -> str:
    """Insere espaços entre letras-e-dígitos grudados.
    Resolve casos como 'VlrrefNF40273DEMILSADAA' → 'VlrrefNF 40273 DEMILSADAA'."""
    if not s:
        return ''
    s = str(s)
    s = re.sub(r'([a-zA-ZÀ-ÿ])(\d)', r'\1 \2', s)
    s = re.sub(r'(\d)([a-zA-ZÀ-ÿ])', r'\1 \2', s)
    return s


def tokenize_historico(historico: str) -> list:
    """Tokeniza o histórico e classifica cada token semanticamente.

    Tipos: CNPJ, CPF, DATE, NUMBER, KEYWORD, NF_MARKER, NAME.
    Devolve lista de tuplas (kind, token_original, valor_processado).

    Pós-processamento: NF_MARKER só permanece se for seguido (eventualmente)
    por NUMBER. Caso contrário vira NAME (resolve 'F C SERVICOS' onde 'F' é
    parte do nome, não marcador).
    """
    if not historico:
        return []
    preprocessed = _preprocess_historico(historico)
    raw_tokens = re.findall(r'\S+', preprocessed)
    out = []
    for tok in raw_tokens:
        bare = re.sub(r'^[^\w]+|[^\w]+$', '', tok)
        if not bare:
            continue
        upper_norm = _normalize_for_keyword(bare)
        # CNPJ formatado
        if re.match(r'^\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}$', tok):
            out.append(('CNPJ', tok, bare))
        # CNPJ raw (14 dígitos)
        elif bare.isdigit() and len(bare) == 14:
            out.append(('CNPJ', tok, bare))
        # CPF formatado
        elif re.match(r'^\d{3}\.\d{3}\.\d{3}-\d{2}$', tok):
            out.append(('CPF', tok, bare))
        # CPF raw (11 dígitos)
        elif bare.isdigit() and len(bare) == 11:
            out.append(('CPF', tok, bare))
        # Data DD/MM/YY[YY]
        elif re.match(r'^\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}$', tok):
            out.append(('DATE', tok, bare))
        # Data curta DD/MM (sem ano) — só se ambos forem 1-31
        elif (m := re.match(r'^(\d{1,2})/(\d{1,2})$', tok)) and 1 <= int(m.group(1)) <= 31 and 1 <= int(m.group(2)) <= 31:
            out.append(('DATE', tok, bare))
        # Valor monetário (X.XXX,XX ou X,XXX.XX) — dígitos com vírgula e/ou ponto
        elif re.match(r'^\$?\d{1,3}([.,]\d{3})*[.,]\d{2}$', tok) or re.match(r'^\d+[.,]\d{2,3}$', tok):
            out.append(('MONEY', tok, bare))
        # Token "R$" sozinho ou junto com valor
        elif upper_norm in ('R', 'RS', 'R$') and re.match(r'^R\$?\d*[.,]?\d*$', tok.upper()):
            out.append(('MONEY', tok, bare))
        # Número puro
        elif bare.isdigit():
            out.append(('NUMBER', tok, bare))
        # Keyword fiscal
        elif upper_norm in FISCAL_KEYWORDS:
            out.append(('KEYWORD', tok, upper_norm))
        # Letra(s) isolada(s) — provável marcador NF ("N F 12345")
        elif upper_norm in NF_MARKERS_SINGLE:
            out.append(('NF_MARKER', tok, upper_norm))
        else:
            out.append(('NAME', tok, bare))

    # Pós-processamento: NF_MARKER sem NUMBER subsequente vira NAME
    # (resolve 'F C SERVICOS' onde 'F' é parte do nome, não marcador)
    for i in range(len(out)):
        if out[i][0] != 'NF_MARKER':
            continue
        is_real_marker = False
        for j in range(i + 1, len(out)):
            nk = out[j][0]
            if nk == 'NF_MARKER':
                continue  # outro marker em sequência, continua olhando
            if nk == 'NUMBER':
                nv = out[j][2]
                if 3 <= len(nv) <= 10:
                    is_real_marker = True
            break  # para no primeiro token não-NF_MARKER
        if not is_real_marker:
            out[i] = ('NAME', out[i][1], out[i][2])
    return out


def smart_extract(historico, col15_raw=None, strategy=0) -> dict:
    """Extrai NF + nome do fornecedor do Complemento Histórico.

    `strategy` escolhe COMO ler o histórico (a usuária confirma o padrão certo na
    tela antes de processar; cada clique em "Repensar" avança pra próxima leitura):
      0 — Tokenização inteligente (padrão)
      1 — Nome após "de/da/do" (ex: 'NFSE nr 3607 de CLINISEG...')
      2 — Nome antes do primeiro número
      3 — Nome = maior trecho de texto (ignora posição)
      4 — NF da coluna do arquivo + nome = histórico limpo
    Retorna sempre: {'value', 'status', 'candidates', 'name'}.
    """
    fn = EXTRACTION_STRATEGIES[strategy % len(EXTRACTION_STRATEGIES)]['fn']
    return fn(historico, col15_raw)


def _strat_tokenizacao(historico, col15_raw=None) -> dict:
    """Estratégia 0 — classificação semântica dos tokens. NF = primeiro NUMBER
    plausível (3–10 díg), bônus após KEYWORD/NF_MARKER; nome = tokens NAME."""
    classified = tokenize_historico(historico)
    if not classified and not col15_raw:
        return {'value': None, 'status': 'EMPTY', 'candidates': [], 'name': ''}

    # Coleta candidatos a NF com score por contexto
    nf_candidates = []  # (valor_str, score, posicao)
    for i, (kind, _raw, val) in enumerate(classified):
        if kind != 'NUMBER':
            continue
        n = len(val)
        if n > 10:
            continue  # >10 = CPF/CNPJ/conta
        prev_is_marker = i > 0 and classified[i - 1][0] in ('NF_MARKER', 'KEYWORD')
        # NF de 2 dígitos só é aceita se vier logo após marker fiscal
        # (evita confundir com mês/dia em texto solto)
        if n < 3 and not prev_is_marker:
            continue
        score = 1
        if prev_is_marker:
            score = 100
        elif i == 0:
            score = 50  # número no INÍCIO do histórico (caso "25443 ALLFOR...")
        nf_candidates.append((str(int(val)), score, i))

    nf_value = None
    nf_status = 'NOT_FOUND'
    candidates_list = []
    if nf_candidates:
        nf_candidates.sort(key=lambda x: (-x[1], x[2]))
        nf_value = nf_candidates[0][0]
        top_score = nf_candidates[0][1]
        same_score_vals = list({c[0] for c in nf_candidates if c[1] == top_score})
        candidates_list = list({c[0] for c in nf_candidates})
        nf_status = 'CONFIDENT' if len(same_score_vals) == 1 else 'AMBIGUOUS'
        if nf_status == 'AMBIGUOUS':
            nf_value = None  # não escolhe automaticamente em caso de empate

    # Fallback / desempate pela col 15
    c15 = normalize_nf(col15_raw)
    if not nf_value and c15:
        if not candidates_list:
            nf_value = c15
            nf_status = 'FROM_COL15'
            candidates_list = [c15]
        elif c15 in candidates_list:
            nf_value = c15
            nf_status = 'AMBIGUOUS_RESOLVED'

    # Razão Social = junção de TODOS os tokens NAME, com limpeza das pontas.
    # Removemos conectores/preposições do INÍCIO e FIM (ex: "de CLINISEG ..."
    # vira "CLINISEG ..."; "... LTDA de" vira "... LTDA"). No meio permanecem
    # ("SERVICOS DE APOIO"). Isso dá uma razão social limpa, melhorando o match
    # por nome no mapa SPED e a busca de CNPJ.
    name_tokens = [val for kind, _raw, val in classified if kind == 'NAME']
    name_tokens = _strip_edge_stopwords(name_tokens)
    name = ' '.join(name_tokens).strip()

    return {
        'value': nf_value,
        'status': nf_status,
        'candidates': candidates_list,
        'name': name,
    }


def _strip_edge_stopwords(tokens: list) -> list:
    """Remove preposições/conectores das PONTAS da lista de tokens do nome.
    No miolo eles permanecem (uma razão social pode conter 'DE', 'E', etc.)."""
    toks = list(tokens)
    while toks and _normalize_for_keyword(toks[0]) in NAME_EDGE_STOPWORDS:
        toks.pop(0)
    while toks and _normalize_for_keyword(toks[-1]) in NAME_EDGE_STOPWORDS:
        toks.pop()
    return toks


def _primeiro_numero(classified):
    """Primeiro NUMBER plausível (3–10 dígitos) da lista classificada. (str|None, pos)."""
    for i, (kind, _raw, val) in enumerate(classified):
        if kind == 'NUMBER' and 3 <= len(val) <= 10:
            return str(int(val)), i
    return None, -1


def _nf_dict(nf_value, name, col15_raw):
    """Empacota o resultado de uma estratégia (com fallback de NF pela col 15)."""
    candidates = [nf_value] if nf_value else []
    status = 'CONFIDENT' if nf_value else 'NOT_FOUND'
    if not nf_value:
        c15 = normalize_nf(col15_raw)
        if c15:
            nf_value, status, candidates = c15, 'FROM_COL15', [c15]
    return {'value': nf_value, 'status': status, 'candidates': candidates, 'name': (name or '').strip()}


_CONECTORES_NOME = (' DE ', ' DA ', ' DO ', ' DOS ', ' DAS ')


def _strat_apos_conector(historico, col15_raw=None) -> dict:
    """Estratégia 1 — nome = tudo APÓS o 1º conector 'de/da/do' (ex: 'NFSE nr 3607
    de CLINISEG SERVICOS...'). NF = 1º número antes do conector (ou o 1º número)."""
    classified = tokenize_historico(historico)
    if not classified and not col15_raw:
        return {'value': None, 'status': 'EMPTY', 'candidates': [], 'name': ''}
    nf_value, _ = _primeiro_numero(classified)
    s = ' ' + _preprocess_historico(str(historico or '')) + ' '
    s_norm = ' ' + _normalize_for_keyword(str(historico or '')) + ' '
    pos = -1
    for con in _CONECTORES_NOME:
        p = s_norm.find(con)
        if p != -1 and (pos == -1 or p < pos):
            pos = p + len(con)
    if pos != -1:
        nome_raw = s[pos:]
        toks = [t for t in re.findall(r'\S+', nome_raw)
                if not re.fullmatch(r'[\d\W]+', t)]
        toks = _strip_edge_stopwords([re.sub(r'^[^\w]+|[^\w]+$', '', t) for t in toks])
        name = ' '.join(t for t in toks if t)
    else:
        name = ' '.join(_strip_edge_stopwords(
            [v for k, _r, v in classified if k == 'NAME']))
    return _nf_dict(nf_value, name, col15_raw)


def _strat_antes_numero(historico, col15_raw=None) -> dict:
    """Estratégia 2 — nome = texto ANTES do 1º número; NF = esse 1º número."""
    classified = tokenize_historico(historico)
    if not classified and not col15_raw:
        return {'value': None, 'status': 'EMPTY', 'candidates': [], 'name': ''}
    nf_value, pos = _primeiro_numero(classified)
    antes = classified[:pos] if pos >= 0 else classified
    toks = [v for k, _r, v in antes if k in ('NAME', 'NF_MARKER')]
    name = ' '.join(_strip_edge_stopwords(toks))
    if not name:  # nada antes do número → cai pros NAME gerais
        name = ' '.join(_strip_edge_stopwords(
            [v for k, _r, v in classified if k == 'NAME']))
    return _nf_dict(nf_value, name, col15_raw)


def _strat_maior_texto(historico, col15_raw=None) -> dict:
    """Estratégia 3 — nome = TODOS os tokens de texto (qualquer posição); NF = o
    número de mais dígitos (3–10)."""
    classified = tokenize_historico(historico)
    if not classified and not col15_raw:
        return {'value': None, 'status': 'EMPTY', 'candidates': [], 'name': ''}
    nums = [(str(int(v)), len(v)) for k, _r, v in classified
            if k == 'NUMBER' and 3 <= len(v) <= 10]
    nf_value = max(nums, key=lambda x: x[1])[0] if nums else None
    name = ' '.join(_strip_edge_stopwords(
        [v for k, _r, v in classified if k == 'NAME']))
    return _nf_dict(nf_value, name, col15_raw)


def _strat_col_nf(historico, col15_raw=None) -> dict:
    """Estratégia 4 — NF vem da COLUNA própria do arquivo (col 15) quando existe;
    nome = histórico limpo (remove só números/datas/CNPJ/marcadores fiscais)."""
    classified = tokenize_historico(historico)
    c15 = normalize_nf(col15_raw)
    nf_first, _ = _primeiro_numero(classified)
    nf_value = c15 or nf_first
    name = ' '.join(_strip_edge_stopwords(
        [v for k, _r, v in classified if k == 'NAME']))
    status = 'FROM_COL15' if c15 else ('CONFIDENT' if nf_value else 'NOT_FOUND')
    return {'value': nf_value, 'status': status,
            'candidates': [nf_value] if nf_value else [], 'name': name.strip()}


def _strat_ancora_empresa(historico, col15_raw=None) -> dict:
    """Estratégia 5 — a razão social começa no TIPO de entidade (FUNDO, BANCO,
    CONDOMINIO, ASSOCIACAO, COOPERATIVA, INDUSTRIA, COMERCIAL…) e vai até o fim.
    Ideal pra históricos com prefixo de descrição: 'DESPESAS CFE AVISO 5359-D
    FUNDO DE INVESTIMENTO...' → nome = 'FUNDO DE INVESTIMENTO...'."""
    classified = tokenize_historico(historico)
    if not classified and not col15_raw:
        return {'value': None, 'status': 'EMPTY', 'candidates': [], 'name': ''}
    nf_value, _ = _primeiro_numero(classified)
    start = None
    for i, (k, _r, v) in enumerate(classified):
        if k == 'NAME' and _normalize_for_keyword(v) in ANCORA_EMPRESA:
            start = i
            break
    if start is not None:
        toks = [v for k, _r, v in classified[start:] if k == 'NAME']
    else:
        toks = [v for k, _r, v in classified if k == 'NAME']
    name = ' '.join(_strip_edge_stopwords(toks))
    return _nf_dict(nf_value, name, col15_raw)


def _strat_apos_ultimo_numero(historico, col15_raw=None) -> dict:
    """Estratégia 6 — a razão social é o texto DEPOIS do último número/código de
    documento (ex: '...AVISO 5359-D FUNDO DE INVESTIMENTO...' → 'FUNDO DE
    INVESTIMENTO...'). Bom quando o nome vem no FIM, após dados do documento."""
    s = _preprocess_historico(str(historico or ''))
    raw = re.findall(r'\S+', s)
    last_digit = -1
    for i, t in enumerate(raw):
        if any(ch.isdigit() for ch in t):
            last_digit = i
    resto = raw[last_digit + 1:] if last_digit >= 0 else raw
    toks = []
    for t in resto:
        bare = re.sub(r'^[^\w]+|[^\w]+$', '', t)
        if not bare or bare.isdigit():
            continue
        un = _normalize_for_keyword(bare)
        if un in FISCAL_KEYWORDS or un in NF_MARKERS_SINGLE:
            continue
        toks.append(bare)
    name = ' '.join(_strip_edge_stopwords(toks))
    if not name:  # nada depois do número → cai pros NAME gerais
        classified = tokenize_historico(historico)
        name = ' '.join(_strip_edge_stopwords([v for k, _r, v in classified if k == 'NAME']))
    nf_value, _ = _primeiro_numero(tokenize_historico(historico))
    return _nf_dict(nf_value, name, col15_raw)


# Registro das estratégias de leitura do histórico (a tela cicla entre elas)
EXTRACTION_STRATEGIES = [
    {'key': 'tokenizacao', 'label': 'Tokenização inteligente (padrão)',
     'descricao': 'Lê o histórico por significado: a NF é o número fiscal mais provável '
                  'e a razão social é o trecho de texto (sem palavras fiscais).',
     'fn': _strat_tokenizacao},
    {'key': 'apos_conector', 'label': 'Nome depois de "de/da/do"',
     'descricao': 'A razão social é o que vem DEPOIS do "de" (ex.: "NFSE nr 3607 de '
                  'CLINISEG SERVICOS..." → nome = CLINISEG SERVICOS...).',
     'fn': _strat_apos_conector},
    {'key': 'antes_numero', 'label': 'Nome antes do número',
     'descricao': 'A razão social é o texto que vem ANTES do número da nota '
                  '(ex.: "CLINISEG SERVICOS LTDA 3607").',
     'fn': _strat_antes_numero},
    {'key': 'maior_texto', 'label': 'Nome = maior trecho de texto',
     'descricao': 'Junta todo o texto do histórico como razão social (ignora a posição) '
                  'e usa o número de mais dígitos como NF.',
     'fn': _strat_maior_texto},
    {'key': 'col_nf', 'label': 'NF da coluna do arquivo + nome limpo',
     'descricao': 'Usa o número da coluna de NF da própria Razão e monta o nome com o '
                  'texto restante do histórico.',
     'fn': _strat_col_nf},
    {'key': 'ancora_empresa', 'label': 'Nome a partir do tipo de empresa',
     'descricao': 'A razão social começa no tipo da entidade (FUNDO, BANCO, CONDOMÍNIO, '
                  'ASSOCIAÇÃO, INDÚSTRIA, COMERCIAL…) e vai até o fim. Ex.: "DESPESAS CFE '
                  'AVISO 5359-D FUNDO DE INVESTIMENTO..." → "FUNDO DE INVESTIMENTO...".',
     'fn': _strat_ancora_empresa},
    {'key': 'apos_ultimo_numero', 'label': 'Nome depois do último número/código',
     'descricao': 'A razão social é o texto que vem DEPOIS do último número ou código de '
                  'documento do histórico (quando o nome vem no fim do texto).',
     'fn': _strat_apos_ultimo_numero},
]


# ============================================================
# APRENDER COM EXEMPLO — a usuária cola um histórico + a razão social correta,
# e a ferramenta deriva o padrão e aplica em toda a base.
# ============================================================
def _tokens_norm(texto):
    """Lista de (índice_raw, palavra_norm) dos tokens com conteúdo de um texto."""
    raw = re.findall(r'\S+', _preprocess_historico(str(texto or '')))
    out = []
    for i, w in enumerate(raw):
        n = _normalize_for_keyword(re.sub(r'^[^\w]+|[^\w]+$', '', w))
        if n:
            out.append((i, n))
    return raw, out


def _find_subseq(hay, needle):
    """Índice (na lista hay) onde a sub-sequência needle começa; -1 se não houver."""
    if not needle or len(needle) > len(hay):
        return -1
    for i in range(len(hay) - len(needle) + 1):
        if hay[i:i + len(needle)] == needle:
            return i
    return -1


def _aprender_regra(historico_ex, razao_ex):
    """Deriva uma regra a partir do exemplo (histórico + razão social correta).
    Identifica ONDE a razão começa e qual o DELIMITADOR antes dela (número, conector
    ou palavra), pra reaplicar em históricos parecidos. Retorna dict ou None."""
    raw, toks = _tokens_norm(historico_ex)
    hnorm = [t[1] for t in toks]
    rnorm = [w for w in re.findall(r'[A-Z0-9]+', _normalize_for_keyword(str(razao_ex))) if w]
    if not rnorm:
        return None
    start = _find_subseq(hnorm, rnorm)
    if start < 0:
        return None  # exemplo não bate com o histórico colado
    to_end = (start + len(rnorm)) >= len(hnorm)
    before = raw[toks[start - 1][0]] if start > 0 else ''
    if before and any(c.isdigit() for c in before):
        delim = ('after_last_number',)
    else:
        bn = _normalize_for_keyword(re.sub(r'^[^\w]+|[^\w]+$', '', before)) if before else ''
        if bn in NAME_EDGE_STOPWORDS:
            delim = ('after_connector', bn)
        elif bn:
            delim = ('after_word', bn)
        else:
            delim = ('from_start',)
    return {'to_end': to_end, 'delim': delim, 'n_palavras': len(rnorm)}


def _descreve_regra(regra):
    d = regra['delim'][0]
    fim = ' até o fim' if regra.get('to_end') else ''
    if d == 'after_last_number':
        return f'A razão social vem depois do último número/código do histórico{fim}.'
    if d == 'after_connector':
        return f'A razão social vem depois de "{regra["delim"][1].lower()}"{fim}.'
    if d == 'after_word':
        return f'A razão social vem depois de "{regra["delim"][1]}"{fim}.'
    return f'A razão social é o texto do histórico (sem palavras fiscais){fim}.'


def _extract_aprendida(historico, col15_raw, regra) -> dict:
    """Aplica a regra aprendida do exemplo a um histórico qualquer."""
    raw = re.findall(r'\S+', _preprocess_historico(str(historico or '')))
    delim = regra.get('delim', ('from_start',))
    inicio = 0
    if delim[0] == 'after_last_number':
        for i, w in enumerate(raw):
            if any(c.isdigit() for c in w):
                inicio = i + 1
    elif delim[0] in ('after_connector', 'after_word'):
        alvo = delim[1]
        for i, w in enumerate(raw):  # primeira ocorrência
            if _normalize_for_keyword(re.sub(r'^[^\w]+|[^\w]+$', '', w)) == alvo:
                inicio = i + 1
                break
    cauda = raw[inicio:]
    toks = []
    for w in cauda:
        bare = re.sub(r'^[^\w]+|[^\w]+$', '', w)
        if not bare or bare.isdigit():
            continue
        un = _normalize_for_keyword(bare)
        if un in FISCAL_KEYWORDS or un in NF_MARKERS_SINGLE:
            continue
        toks.append(bare)
    name = ' '.join(_strip_edge_stopwords(toks))
    nf_value, _ = _primeiro_numero(tokenize_historico(historico))
    return _nf_dict(nf_value, name, col15_raw)


def montar_extrator(estrategia=0, exemplo_historico=None, exemplo_razao=None):
    """Decide COMO extrair e devolve (extrator_fn, info). Se houver exemplo:
    1) procura uma das leituras prontas que reproduz o exemplo; senão
    2) deriva uma regra própria do exemplo. Sem exemplo, usa a `estrategia`."""
    if exemplo_historico and exemplo_razao and str(exemplo_razao).strip():
        alvo = ''.join(re.findall(r'[A-Z0-9]+', _normalize_for_keyword(str(exemplo_razao))))
        # 1) alguma leitura pronta reproduz o exemplo?
        for i, s in enumerate(EXTRACTION_STRATEGIES):
            got = s['fn'](exemplo_historico, None).get('name', '')
            if alvo and ''.join(re.findall(r'[A-Z0-9]+', _normalize_for_keyword(got))) == alvo:
                fn = s['fn']
                return ((lambda h, c: fn(h, c)),
                        {'idx': i, 'label': f'Aprendido do seu exemplo · leitura "{s["label"]}"',
                         'descricao': f'Seu exemplo casou com a leitura "{s["label"]}". '
                                      f'Aplicando esse padrão em toda a base.'})
        # 2) deriva regra própria
        regra = _aprender_regra(exemplo_historico, exemplo_razao)
        if regra:
            return ((lambda h, c: _extract_aprendida(h, c, regra)),
                    {'idx': -1, 'label': 'Aprendido do seu exemplo · regra própria',
                     'descricao': _descreve_regra(regra)})
        # 3) exemplo não pôde ser aprendido → segue com a estratégia escolhida
    s = EXTRACTION_STRATEGIES[estrategia % len(EXTRACTION_STRATEGIES)]
    fn = s['fn']
    return ((lambda h, c: fn(h, c)),
            {'idx': estrategia % len(EXTRACTION_STRATEGIES),
             'label': s['label'], 'descricao': s['descricao']})


def extract_nf_from_historico(complemento, col15_raw=None, strategy=0) -> dict:
    """Extrai NF do Complemento Histórico. Wrapper sobre smart_extract."""
    r = smart_extract(complemento, col15_raw=col15_raw, strategy=strategy)
    return {
        'value': r['value'],
        'status': r['status'],
        'candidates': r['candidates'],
    }


def extract_cnpj_from_historico(complemento) -> str | None:
    """Se o histórico já cita o CNPJ (formatado ou 14 dígitos), extrai direto.
    Mais confiável que tentar adivinhar pelo nome do fornecedor."""
    if not complemento or (isinstance(complemento, float) and pd.isna(complemento)):
        return None
    s = str(complemento)
    m = RE_CNPJ_FORMATADO.search(s)
    if m:
        return normalize_cnpj(m.group(1))
    m = RE_CNPJ_RAW.search(s)
    if m:
        return normalize_cnpj(m.group(1))
    return None


def extract_supplier_name(complemento, strategy=0) -> str:
    """Extrai a Razão Social do Complemento Histórico (estratégia escolhida)."""
    return smart_extract(complemento, strategy=strategy)['name']


def preview_extracao(df_razao, strategy=0, exemplo_historico=None, exemplo_razao=None,
                     limite=25) -> dict:
    """Roda a extração (NF + Razão Social) numa AMOSTRA de linhas distintas da Razão
    pra a usuária conferir o padrão ANTES de processar tudo. Se houver exemplo, usa o
    padrão aprendido dele. Devolve o rótulo da leitura + exemplos {historico, nf, razao}."""
    col_hist = 13  # Complemento Histórico (col 14, 0-indexed 13)
    col_nf15 = 14  # NF original (col 15)
    extrator, info = montar_extrator(strategy, exemplo_historico, exemplo_razao)
    vistos = set()
    amostras = []
    for row in df_razao.itertuples(index=False):
        hist = row[col_hist] if len(row) > col_hist else None
        if hist is None or (isinstance(hist, float) and pd.isna(hist)):
            continue
        hist_s = str(hist).strip()
        if not hist_s:
            continue
        chave = hist_s[:60].upper()
        if chave in vistos:
            continue
        vistos.add(chave)
        col15 = row[col_nf15] if len(row) > col_nf15 else None
        r = extrator(hist_s, col15)
        amostras.append({
            'historico': hist_s,
            'nf': r['value'] or '—',
            'razao': r['name'] or '—',
        })
        if len(amostras) >= limite:
            break
    return {
        'estrategia_idx': info['idx'],
        'estrategia_label': info['label'],
        'estrategia_descricao': info['descricao'],
        'total_estrategias': len(EXTRACTION_STRATEGIES),
        'aprendido': info['idx'] == -1 or 'Aprendido' in info['label'],
        'amostras': amostras,
    }


def build_nf_only_index(agg_nf_map: dict) -> dict:
    """Constrói um índice <NF> → lista de (CNPJ, entry) a partir do agg_map de
    C100-EFD/A100 (cujas chaves são "<NF>|<CNPJ>"). Permite buscar uma NF nos
    blocos sem precisar do CNPJ — usado como último fallback quando a Razão
    não dá pistas suficientes pra resolver o CNPJ pelo nome do fornecedor.
    """
    idx: dict[str, list] = {}
    for key, entry in agg_nf_map.items():
        if '|' not in key:
            continue
        nf, cnpj = key.split('|', 1)
        idx.setdefault(nf, []).append((cnpj, entry))
    return idx


def lookup_cnpj_by_nf_in_speds(nf_value, vlr_partida, nf_idx_list) -> dict | None:
    """Quarta camada de fallback: dado um número de NF e uma lista de índices
    NF-only (C100-EFD, A100 com crédito, A100 sem crédito, etc.), tenta achar
    qual CNPJ tem essa NF nos próprios SPEDs.

      • Se 1 único CNPJ tem essa NF em algum bloco → usa direto (via NF_SPED_UNICO).
      • Se múltiplos CNPJs tem → usa Vlr Partida como desempate (mais próximo
        dentro da tolerância → via NF_SPED_VALOR).
      • Caso contrário → None (não é seguro inferir CNPJ).

    nf_idx_list: lista de dicts construídos por build_nf_only_index, em ordem
    de prioridade (C100-EFD primeiro, A100 depois).
    """
    if not nf_value or not nf_idx_list:
        return None

    for idx in nf_idx_list:
        candidates = idx.get(nf_value)
        if not candidates:
            continue
        # Dedup CNPJs (uma mesma NF+CNPJ não aparece duas vezes no mesmo idx,
        # mas se aparecer em vários idx, o loop externo trata cada um)
        if len(candidates) == 1:
            cnpj_found, _entry = candidates[0]
            return {'cnpj': cnpj_found, 'via': 'NF_SPED_UNICO'}
        # Múltiplos → tenta desambiguar pelo valor mais próximo da Vlr Partida
        best_cnpj, best_entry = min(
            candidates, key=lambda c: abs(c[1]['total'] - vlr_partida)
        )
        if abs(best_entry['total'] - vlr_partida) <= TOLERANCIA_VALOR:
            return {'cnpj': best_cnpj, 'via': 'NF_SPED_VALOR'}
        # Ambíguo: vários CNPJs com essa NF e nenhum bate com Vlr Partida.
        # Não retorna nada — preferimos deixar sem CNPJ a inferir errado.
    return None


# Cache de busca web — chave: nome_normalizado → {cnpj, razao_social} ou None (falhou)
NAME_CNPJ_WEB_CACHE: dict = {}
WEB_SEARCH_THROTTLE = 3.0  # segundos entre buscas (educado com APIs anti-bot)
WEB_SEARCH_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
]
# Contador de falhas consecutivas por fonte — quando atinge limite, pula a fonte
_SOURCE_FAILURES: dict = {'bing': 0, 'ddg': 0, 'cnpjbiz': 0}
_SOURCE_BLOCKED_THRESHOLD = 5  # se 5 buscas seguidas falharem → fonte bloqueada


def _pick_user_agent() -> str:
    import random
    return random.choice(WEB_SEARCH_USER_AGENTS)


def _extract_cnpjs_from_html(html: str) -> list[str]:
    """Extrai todos os CNPJs (formatados ou raw 14 dígitos) de um HTML."""
    found = []
    seen = set()
    for m in re.finditer(r'\b(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})\b', html):
        c = normalize_cnpj(m.group(1))
        if len(c) == 14 and c not in seen:
            seen.add(c)
            found.append(c)
    return found


def _search_bing(query: str) -> list[str]:
    """Busca no Bing e devolve candidatos a CNPJ encontrados no HTML."""
    headers = {
        'User-Agent': _pick_user_agent(),
        'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
        'Accept': 'text/html,application/xhtml+xml',
    }
    try:
        r = requests.get(
            'https://www.bing.com/search',
            params={'q': query, 'cc': 'BR', 'setlang': 'pt-BR'},
            headers=headers,
            timeout=12,
        )
        if not r.ok:
            return []
        return _extract_cnpjs_from_html(r.text)
    except Exception as e:
        log.debug('Bing falhou: %s', e)
        return []


def _search_ddg(query: str) -> list[str]:
    """Busca no DuckDuckGo HTML e devolve candidatos a CNPJ."""
    headers = {
        'User-Agent': _pick_user_agent(),
        'Accept-Language': 'pt-BR,pt;q=0.9',
        'Accept': 'text/html,application/xhtml+xml',
    }
    try:
        r = requests.post(
            'https://html.duckduckgo.com/html/',
            data={'q': query},
            headers=headers,
            timeout=12,
        )
        if not r.ok:
            return []
        return _extract_cnpjs_from_html(r.text)
    except Exception as e:
        log.debug('DDG falhou: %s', e)
        return []


def _search_cnpjbiz(name: str) -> list[str]:
    """Busca no cnpj.biz (site especializado em CNPJ) por nome."""
    headers = {
        'User-Agent': _pick_user_agent(),
        'Accept-Language': 'pt-BR,pt;q=0.9',
    }
    try:
        r = requests.get(
            'https://cnpj.biz/procura',
            params={'q': name},
            headers=headers,
            timeout=12,
        )
        if not r.ok:
            return []
        return _extract_cnpjs_from_html(r.text)
    except Exception as e:
        log.debug('cnpj.biz falhou: %s', e)
        return []


def _all_sources_blocked() -> bool:
    """True quando TODAS as fontes web já desistiram (bloqueio anti-bot).
    Em servidores (IP de datacenter, ex: Render), as 3 fontes caem rápido —
    aí não adianta continuar tentando/dormindo: a busca web é abortada."""
    return all(_SOURCE_FAILURES.get(k, 0) >= _SOURCE_BLOCKED_THRESHOLD
               for k in ('bing', 'ddg', 'cnpjbiz'))


def _try_source(src_key: str, search_fn, query: str) -> list[str]:
    """Tenta uma fonte de busca. Se ela está marcada como bloqueada, pula."""
    if _SOURCE_FAILURES.get(src_key, 0) >= _SOURCE_BLOCKED_THRESHOLD:
        return []  # fonte já desistiu pra essa execução
    candidates = search_fn(query)
    if candidates:
        _SOURCE_FAILURES[src_key] = 0  # sucesso reseta contador
    else:
        _SOURCE_FAILURES[src_key] = _SOURCE_FAILURES.get(src_key, 0) + 1
        if _SOURCE_FAILURES[src_key] == _SOURCE_BLOCKED_THRESHOLD:
            log.warning('Fonte web "%s" parece bloqueada (5 falhas seguidas) — pulando.', src_key)
    return candidates


def search_cnpj_by_name_web(supplier_name: str) -> dict | None:
    """5ª camada: descobre o CNPJ pesquisando o nome do fornecedor na web.

    Tenta 3 fontes em sequência (Bing → DuckDuckGo → cnpj.biz). Cada fonte é
    descartada após 5 falhas consecutivas (provável bloqueio anti-bot).
    Todos os candidatos coletados são validados via BrasilAPI/CNPJa: só aceita
    o primeiro CNPJ cujo nome oficial é similar (Levenshtein ≥ 0.55) ao buscado.
    """
    if not supplier_name or len(supplier_name) < 8:
        return None
    key = normalize_name(supplier_name)
    if not key:
        return None
    if key in NAME_CNPJ_WEB_CACHE:
        return NAME_CNPJ_WEB_CACHE[key]

    query = f'{supplier_name} CNPJ'
    # Coleta candidatos de todas as fontes disponíveis
    candidates: list[str] = []
    for src_key, src_fn in (('bing', _search_bing),
                            ('ddg', _search_ddg),
                            ('cnpjbiz', lambda q: _search_cnpjbiz(supplier_name))):
        cands = _try_source(src_key, src_fn, query)
        for c in cands:
            if c not in candidates:
                candidates.append(c)
        # Se já temos 5+ candidatos, não precisa consultar mais fontes
        if len(candidates) >= 5:
            break

    if not candidates:
        NAME_CNPJ_WEB_CACHE[key] = None
        return None

    # Valida candidatos via BrasilAPI/CNPJa — só aceita se o nome bater
    for cnpj in candidates[:8]:
        data = fetch_cnae_once(cnpj)
        if not data or not data.get('razao_social'):
            continue
        # Reconcilia o nome OFICIAL (da Receita) com o nome buscado (do histórico).
        # Limiar alto: a web traz muito candidato parecido-mas-errado.
        sim = _names_reconcile(supplier_name, data['razao_social'])
        if sim >= 0.70:
            result = {
                'cnpj': cnpj,
                'via': 'WEB_SEARCH',
                'matched_name': data['razao_social'],
                'similarity': round(sim, 2),
            }
            NAME_CNPJ_WEB_CACHE[key] = result
            log.info('Web search "%s" → %s (%s, sim=%.2f)',
                     supplier_name[:50], cnpj, data['razao_social'][:40], sim)
            return result
    NAME_CNPJ_WEB_CACHE[key] = None
    return None


def lookup_cnpj_by_historico_scan(complemento, cnpj_map) -> dict | None:
    """Reverse lookup: varre o Complemento Histórico procurando nomes de fornecedores
    conhecidos (do mapa SPED) como substring. Funciona mesmo quando a extração de
    nome falha — basta que o nome do fornecedor esteja em algum lugar do histórico.

    Retorna o match de maior comprimento (mais específico) ou None.
    """
    if not complemento or not cnpj_map:
        return None
    norm_hist = normalize_name(complemento)
    if len(norm_hist) < 8:
        return None
    best_cnpj = None
    best_name = ''
    for sup_name, cnpj in cnpj_map.items():
        if len(sup_name) < 8:
            continue  # nomes muito curtos podem causar falso positivo
        if sup_name in norm_hist and len(sup_name) > len(best_name):
            best_cnpj = cnpj
            best_name = sup_name
    if best_cnpj:
        return {'cnpj': best_cnpj, 'via': 'HISTORICO_SCAN', 'matched_name': best_name}
    return None


# Mantida só para compatibilidade — a lógica agora está dentro de extract_nf_from_historico.
def resolve_nf_extraction(extracted: dict, col15_raw) -> dict:
    return extracted


def lookup_cnpj(extracted_name: str, cnpj_map: dict) -> dict | None:
    """Procura CNPJ a partir do nome extraído. 4 camadas de match em ordem:
      1. EXACT — match exato após normalização
      2. STRIP_DIGITS — remove CPF/CNPJ no fim do nome e tenta de novo
      3. PREFIX — chave do mapa é prefixo do nome extraído (ou vice-versa)
      4. FUZZY — similaridade Levenshtein ≥ 85% (cobre variações como "LTDA" vs "LTDA ME")
    """
    if not extracted_name:
        return None
    norm = normalize_name(extracted_name)
    if not norm:
        return None
    # 1. Exato
    if norm in cnpj_map:
        return {'cnpj': cnpj_map[norm], 'via': 'EXACT'}
    # 2. Strip CPF/CNPJ no fim
    stripped = re.sub(r'\d{11,14}$', '', norm)
    if stripped and stripped != norm and stripped in cnpj_map:
        return {'cnpj': cnpj_map[stripped], 'via': 'STRIP_DIGITS'}
    # 3. Prefixo (cobre abreviações)
    if len(stripped) >= 10:
        for k, v in cnpj_map.items():
            if k == stripped:
                return {'cnpj': v, 'via': 'EXACT_STRIPPED'}
            if (k.startswith(stripped) or stripped.startswith(k)) and min(len(k), len(stripped)) >= 10:
                return {'cnpj': v, 'via': 'PREFIX', 'matched_name': k}
    # 4. Fuzzy match (Levenshtein) — cobre "EMPRESA LTDA" vs "EMPRESA LTDA ME", etc.
    # Só dispara pra nomes longos o suficiente pra evitar falso positivo
    needle = stripped if len(stripped) >= 10 else norm
    if len(needle) >= 10 and cnpj_map:
        matches = difflib.get_close_matches(needle, list(cnpj_map.keys()), n=1, cutoff=0.85)
        if matches:
            return {'cnpj': cnpj_map[matches[0]], 'via': 'FUZZY', 'matched_name': matches[0]}
    return None


# ============================================================
# CNAE LOOKUP — BrasilAPI primário + CNPJa fallback, com cache
# ============================================================
def _is_valid_result(d: dict) -> bool:
    """Considera resultado válido só quando tem razão social ou CNAE preenchido."""
    return bool(d) and bool(d.get('razao_social') or d.get('cnae1_desc'))


def _parse_brasilapi(d: dict) -> dict:
    result = {
        'razao_social': d.get('razao_social') or d.get('nome_fantasia') or '',
        'cnae1_desc': d.get('cnae_fiscal_descricao') or '',
        'cnae2_desc': '',
        'source': 'BrasilAPI',
    }
    sec = d.get('cnaes_secundarios') or []
    if sec and isinstance(sec, list) and sec[0]:
        result['cnae2_desc'] = sec[0].get('descricao') or ''
    return result


def _parse_cnpja(d: dict) -> dict:
    main = d.get('mainActivity') or {}
    sides = d.get('sideActivities') or []
    side = sides[0] if sides else {}
    company = d.get('company') or {}
    return {
        'razao_social': company.get('name') or d.get('alias') or '',
        'cnae1_desc': main.get('text') or '',
        'cnae2_desc': side.get('text') or '',
        'source': 'CNPJa',
    }


def _parse_receitaws(d: dict) -> dict | None:
    if d.get('status') == 'ERROR':
        return None
    main = d.get('atividade_principal') or []
    main_first = main[0] if main and isinstance(main, list) else {}
    sec = d.get('atividades_secundarias') or []
    sec_first = sec[0] if sec and isinstance(sec, list) else {}
    return {
        'razao_social': d.get('nome') or d.get('fantasia') or '',
        'cnae1_desc': main_first.get('text') if isinstance(main_first, dict) else '',
        'cnae2_desc': sec_first.get('text') if isinstance(sec_first, dict) else '',
        'source': 'ReceitaWS',
    }


def _try_api(url: str, parse_fn, source_name: str, cnpj: str) -> dict | None:
    """Tenta uma API com retry em HTTP 429 (rate limit). Devolve dict válido ou None."""
    headers = {'User-Agent': USER_AGENT}
    for attempt in range(2):  # 2 tentativas (original + 1 retry após 429)
        try:
            r = requests.get(url, headers=headers, timeout=CNAE_TIMEOUT)
            if r.status_code == 429:
                # Rate limit — espera antes de tentar de novo
                wait = 2 + attempt * 2  # 2s, depois 4s
                log.debug('%s rate limit em %s, aguardando %ss', source_name, cnpj, wait)
                time.sleep(wait)
                continue
            if r.ok:
                try:
                    parsed = parse_fn(r.json())
                except Exception as e:
                    log.debug('%s parse falhou em %s: %s', source_name, cnpj, e)
                    return None
                if parsed and _is_valid_result(parsed):
                    return parsed
                return None
            break  # outro status code — não retry
        except Exception as e:
            log.debug('%s %s tentativa %s falhou: %s', source_name, cnpj, attempt, e)
            time.sleep(1)
    return None


def fetch_cnae_once(cnpj: str) -> dict:
    """Tenta consultar CNPJ em 3 fontes em sequência. Só cacheia sucesso."""
    # Cache só vale se for um resultado real
    cached = CNAE_CACHE.get(cnpj)
    if cached and _is_valid_result(cached):
        return cached

    # 1. BrasilAPI
    result = _try_api(
        f'https://brasilapi.com.br/api/cnpj/v1/{cnpj}',
        _parse_brasilapi, 'BrasilAPI', cnpj,
    )
    # 2. CNPJa
    if not result:
        result = _try_api(
            f'https://open.cnpja.com/office/{cnpj}',
            _parse_cnpja, 'CNPJa', cnpj,
        )
    # 3. ReceitaWS
    if not result:
        result = _try_api(
            f'https://receitaws.com.br/v1/cnpj/{cnpj}',
            _parse_receitaws, 'ReceitaWS', cnpj,
        )

    if result and _is_valid_result(result):
        CNAE_CACHE[cnpj] = result  # só cacheia se for sucesso real
        return result
    return {'razao_social': '', 'cnae1_desc': '', 'cnae2_desc': '', 'source': 'ERROR'}


def fetch_all_cnaes(cnpjs: list[str]) -> dict[str, dict]:
    """Busca paralela em 2 passadas: paralelo + retry sequencial dos que falharam.

    1ª passada: paralelo com CNAE_CONCURRENCY threads, retry em 429 dentro do request.
    2ª passada: sequencial, espaçada (2s entre cada), recupera o que falhou.
    Só cacheia sucessos — falhas são re-tentadas em execuções futuras.
    """
    EMPTY = {'razao_social': '', 'cnae1_desc': '', 'cnae2_desc': '', 'source': 'ERROR'}
    results = {}

    # Pega do cache somente os sucessos (cache não armazena mais falhas, mas defesa em profundidade)
    to_fetch = []
    for c in cnpjs:
        cached = CNAE_CACHE.get(c)
        if cached and _is_valid_result(cached):
            results[c] = cached
        else:
            to_fetch.append(c)
    log.info('Consulta CNPJ: %s já em cache · %s a buscar',
             len(cnpjs) - len(to_fetch), len(to_fetch))
    if not to_fetch:
        return results

    # 1ª PASSADA: paralelo
    with ThreadPoolExecutor(max_workers=CNAE_CONCURRENCY) as ex:
        futures = {ex.submit(fetch_cnae_once, c): c for c in to_fetch}
        for fut in as_completed(futures):
            c = futures[fut]
            try:
                results[c] = fut.result() or dict(EMPTY)
            except Exception as e:
                log.warning('Consulta %s erro: %s', c, e)
                results[c] = dict(EMPTY)

    falhou = [c for c in to_fetch if not _is_valid_result(results.get(c, {}))]
    if not falhou:
        return results

    # 2ª PASSADA: sequencial, espaçada (respeita rate limit)
    log.info('1ª passada: %s CNPJs falharam. Iniciando 2ª passada sequencial…',
             len(falhou))
    for i, c in enumerate(falhou, start=1):
        time.sleep(2)  # 2s entre consultas
        data = fetch_cnae_once(c)
        if _is_valid_result(data):
            results[c] = data
            log.debug('2ª passada %s/%s: %s recuperado via %s',
                      i, len(falhou), c, data.get('source'))
    recuperados = sum(1 for c in falhou if _is_valid_result(results.get(c, {})))
    log.info('2ª passada concluída: %s recuperados · %s ainda falharam',
             recuperados, len(falhou) - recuperados)
    return results


# ============================================================
# LEITURA E AGREGAÇÃO DOS SPEDs
# ============================================================
# Schema dos blocos (1-indexed igual ao usado no front-end)
# Col 0 (1-indexed: 1) sempre é o CNPJ da EMPRESA QUE REPORTA (a empresa-raiz que está sendo auditada).
# Os outros CNPJs (col 'cnpj' aqui) são dos PARTICIPANTES (fornecedores/clientes).
SPED_SCHEMA = {
    'c100efd': {
        'block_key': 'c100efd',
        'required_fields': ['cnpj', 'nome', 'nf', 'valor'],
        'apply_cst': False,
        'kind': 'nf',
    },
    'a100': {
        'block_key': 'a100',
        'required_fields': ['cnpj', 'nome', 'nf', 'cst', 'valor'],
        'apply_cst': True,
        'kind': 'nf',
    },
    'f100': {
        'block_key': 'f100',
        'required_fields': ['cnpj', 'nome', 'vlr_operacao', 'cst', 'valor'],
        'apply_cst': True,
        'kind': 'f100',
    },
}


def _safe_read_excel(file_stream, friendly_name, **kwargs):
    """
    Envolve pd.read_excel num try/except amigável que diz QUAL arquivo falhou
    e por quê (xlsx inválido, arquivo vazio, formato xls antigo etc).
    """
    try:
        if hasattr(file_stream, 'seek'):
            try:
                file_stream.seek(0)
            except Exception:
                pass
        return pd.read_excel(file_stream, engine='openpyxl', **kwargs)
    except Exception as e:
        msg = str(e).lower()
        if 'not a zip file' in msg or 'badzipfile' in msg:
            raise ValueError(
                f'Não foi possível abrir "{friendly_name}". O arquivo não parece ser '
                f'um .xlsx válido (pode estar vazio, corrompido, em formato .xls antigo, '
                f'ou ser um .csv renomeado). Reabra no Excel e salve novamente como '
                f'"Pasta de Trabalho do Excel (*.xlsx)".'
            )
        raise ValueError(f'Erro ao ler "{friendly_name}": {e}')


# Padrões de regex (sem acentos, uppercase) pra detectar colunas SPED pelo NOME do cabeçalho.
# Cada campo lógico tem 1+ regex que tenta casar. Pega o PRIMEIRO match.
# Isso resolve o problema de layouts variados entre exportadores SPED diferentes.
SPED_HEADER_PATTERNS = {
    'c100efd': {
        'cnpj_empresa': [r'^CNPJ$'],
        'periodo': [r'^PERIODO$'],
        'cnpj': [r'CNPJ.*PARTICIPANTE'],
        'nome': [r'NOME.*PARTICIPANTE'],
        'nf': [r'NUMERO.*DOCUMENTO', r'\bNUM\b.*\bDOC\b', r'^NUM\s*DOC'],
        'valor': [r'^VLR\s*ITEM$', r'^VALOR\s*ITEM$'],
    },
    'a100': {
        'cnpj_empresa': [r'^CNPJ$'],
        'periodo': [r'^PERIODO$'],
        'cnpj': [r'CNPJ.*PARTICIPANTE'],
        'nome': [r'NOME.*PARTICIPANTE'],
        'nf': [r'NUMERO.*DOCUMENTO', r'\bNUM\b.*\bDOC\b'],
        # CST Cofins (item, NÃO o A100 header)
        'cst': [r'^CST\s*COFINS$'],
        # Vlr Base Cálculo Cofins (item) — exclui o "Vlr Base Cálculo Cofins - A100" (header)
        'valor': [r'^VLR\s*BASE\s*CALCULO\s*COFINS$'],
    },
    'f100': {
        'cnpj_empresa': [r'^CNPJ$'],
        'periodo': [r'^PERIODO$'],
        'cnpj': [r'CNPJ.*PARTICIPANTE'],
        'nome': [r'NOME.*PARTICIPANTE'],
        'vlr_operacao': [r'^VLR\s*OPERACAO$'],
        'cst': [r'^CST\s*COFINS$'],
        'valor': [r'^VLR\s*BASE\s*CALCULO\s*COFINS$'],
    },
}


def _detect_columns_by_name(headers: list, patterns: dict) -> dict:
    """Mapeia campo_lógico → índice da coluna procurando pelo nome do header.
    Retorna dict {campo: índice_0_based} apenas pros campos encontrados.

    Pra campos onde múltiplas colunas casam (ex: cnpj_empresa casa com 'CNPJ Participante'),
    usa o PRIMEIRO match. Pra cnpj/cnpj_empresa, faz desambiguação especial.
    """
    mapping: dict[str, int] = {}
    norm_headers = [_normalize_for_keyword(str(h)) for h in headers]
    for field, regex_list in patterns.items():
        for idx, h in enumerate(norm_headers):
            if idx in mapping.values() and field != 'cnpj_empresa':
                # já mapeado pra outro campo (cnpj_empresa pode aparecer 2x)
                pass
            for pat in regex_list:
                if re.search(pat, h):
                    if field not in mapping:
                        mapping[field] = idx
                    break
            if field in mapping:
                break
    # Especial: cnpj_empresa é a PRIMEIRA col "CNPJ" (sem "PARTICIPANTE")
    # Re-mapeia se necessário
    for idx, h in enumerate(norm_headers):
        if h == 'CNPJ' and mapping.get('cnpj_empresa') != idx:
            mapping['cnpj_empresa'] = idx
            break
    return mapping


def read_sped(file_stream, schema, friendly_name='SPED'):
    """
    Detecta as colunas de um SPED pelo NOME do cabeçalho (não posição fixa) e
    valida que os campos essenciais existem. Retorna col_map (campo lógico →
    índice 0-based da coluna no arquivo). NÃO lê os dados — a agregação é feita
    em streaming (openpyxl read_only), muito mais rápido que o pandas.
    """
    header_df = _safe_read_excel(
        file_stream, friendly_name=friendly_name, header=0, nrows=0, dtype=object
    )
    headers = list(header_df.columns)
    block_key = schema.get('block_key', 'c100efd')
    patterns = SPED_HEADER_PATTERNS.get(block_key, {})
    col_map = _detect_columns_by_name(headers, patterns)
    log.info('SPED "%s" (%s cols) — colunas detectadas: %s',
             friendly_name, len(headers), {k: f'{v+1} ({headers[v]})' for k, v in col_map.items()})

    required = schema.get('required_fields', list(patterns.keys()))
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(
            f'Não consegui localizar as colunas {missing} em "{friendly_name}". '
            f'Cabeçalhos disponíveis: {", ".join(str(h) for h in headers[:20])}...'
        )
    return col_map


def process_sped_block(file_stream, schema, cnpj_map: dict, friendly_name='SPED'):
    """
    Agrega um SPED lendo o arquivo em STREAMING (openpyxl read_only) — sem pandas
    e sem segurar o arquivo na RAM. Muito mais rápido em arquivos largos (84 colunas
    × dezenas de milhares de linhas) e popula o mapa Nome→CNPJ.

    Para A100/F100 (apply_cst=True): agg (CST 50-67, com crédito) + agg_no_cred (resto).
    Pra C100-EFD (apply_cst=False): só agg.
    """
    from openpyxl import load_workbook as _load_ro
    col_map = read_sped(file_stream, schema, friendly_name=friendly_name)
    apply_cst = schema['apply_cst']
    kind = schema['kind']

    agg_map = {}
    agg_no_cred = {} if apply_cst else None
    skipped = 0
    cnpj_added = 0
    company_cnpj_count = {}
    periodos = set()

    i_ce = col_map.get('cnpj_empresa'); i_cnpj = col_map.get('cnpj')
    i_nome = col_map.get('nome'); i_nf = col_map.get('nf'); i_val = col_map.get('valor')
    i_cst = col_map.get('cst'); i_per = col_map.get('periodo'); i_op = col_map.get('vlr_operacao')

    def g(vals, idx):
        return vals[idx] if (idx is not None and idx < len(vals)) else None

    def _push_nf(target, nf_n, cnpj_val, value):
        entry = target.get(f'{nf_n}|{cnpj_val}')
        if entry:
            entry['total'] += value
            entry['items'] += 1
        else:
            target[f'{nf_n}|{cnpj_val}'] = {'total': value, 'items': 1}

    def _push_f100(target, cnpj_val, per, op, bc):
        key = f'{cnpj_val}|{per}'
        if target.get(key) is None:
            target[key] = [{'op': op, 'bc': bc}]
        else:
            target[key].append({'op': op, 'bc': bc})

    if hasattr(file_stream, 'seek'):
        file_stream.seek(0)
    wb_in = _load_ro(file_stream, read_only=True, data_only=True)
    src = wb_in.active
    rows = src.iter_rows(values_only=True)
    next(rows, None)  # pula cabeçalho
    total = 0
    for vals in rows:
        total += 1
        ce_raw = g(vals, i_ce)
        if ce_raw is not None:
            ce_norm = normalize_cnpj(ce_raw)
            if len(ce_norm) == 14:
                company_cnpj_count[ce_norm] = company_cnpj_count.get(ce_norm, 0) + 1
        if i_per is not None:
            _per = to_month_year(g(vals, i_per))
            if _per:
                periodos.add(_per)

        has_credit = True if not apply_cst else is_credit_cst(g(vals, i_cst))
        target = agg_map if has_credit else agg_no_cred
        if not has_credit:
            skipped += 1
        if target is None:
            continue

        name_key = normalize_name(g(vals, i_nome))
        cnpj_val = normalize_cnpj(g(vals, i_cnpj))
        if name_key and len(cnpj_val) == 14 and name_key not in cnpj_map:
            cnpj_map[name_key] = cnpj_val
            cnpj_added += 1

        if kind == 'nf':
            nf_n = normalize_nf(g(vals, i_nf))
            if not nf_n or not cnpj_val:
                continue
            _push_nf(target, nf_n, cnpj_val, to_number(g(vals, i_val)))
        elif kind == 'f100':
            per = to_month_year(g(vals, i_per))
            if not cnpj_val or not per:
                continue
            _push_f100(target, cnpj_val, per, to_number(g(vals, i_op)), to_number(g(vals, i_val)))
    wb_in.close()

    company_cnpj = (
        max(company_cnpj_count.items(), key=lambda x: x[1])[0]
        if company_cnpj_count else None
    )
    return {
        'agg': agg_map,
        'agg_no_cred': agg_no_cred,
        'skipped': skipped,
        'total': total,
        'cnpj_added': cnpj_added,
        'company_cnpj': company_cnpj,
        'col_map': col_map,    # campo lógico → índice de coluna no arquivo original
        'periodos': periodos,  # períodos (MM/AAAA) presentes no bloco
    }


def find_f100_match(f100_map, cnpj, periodo, vlr_partida):
    """Para F100, procura operações com Vlr Operação ≈ Vlr Partida (tolerância)."""
    key = f'{cnpj}|{periodo}'
    lst = f100_map.get(key)
    if not lst:
        return None
    matches = [x for x in lst if abs(x['op'] - vlr_partida) <= TOLERANCIA_VALOR]
    if matches:
        total = sum(x['bc'] for x in matches)
        return {'total': total, 'count': len(matches), 'exact_match': True}
    # Sem match exato: ainda assim retorna a soma do CNPJ+período (informativo)
    total = sum(x['bc'] for x in lst)
    return {'total': total, 'count': len(lst), 'exact_match': False}


# ============================================================
# ANÁLISE DO CRUZAMENTO
# ============================================================
def _fmt_brl(value):
    """Formata número como BR currency 'R$ 1.234,56'."""
    s = f'{abs(value):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f'R$ {s}'


def build_analise(vlr_partida, val_efd, val_a100, val_f100,
                  partidas_match_efd=False, partidas_match_a100=False,
                  total_partidas=None, partidas_count=1,
                  a100_no_credit=False, f100_no_credit=False,
                  efd_via_nf_only=False):
    """Devolve os 3 campos finais (separados pra facilitar filtros):
      - analise_text / analise_style  → status enxuto na col ANÁLISE
      - soma_partidas_text            → texto da col "Fecha somando partidas" (ou vazio)
      - diferenca_value               → número da col "Valor da diferença" (ou None)

    Quando `a100_no_credit` ou `f100_no_credit` é True, a NF foi localizada
    no bloco mas em CST fora de 50-67 (sem crédito).
    Quando `efd_via_nf_only` é True, a NF foi localizada no C100-EFD por NF only
    (CNPJ não foi confirmado pelo cruzamento estrito).
    """
    # Blocos com match confirmado (NF + CNPJ batendo) e com crédito
    blocos_cred = []
    if val_efd is not None and not efd_via_nf_only:
        blocos_cred.append('C100-EFD')
    if val_a100 is not None and not a100_no_credit:
        blocos_cred.append('A100-CONTRI')
    if val_f100 is not None and not f100_no_credit:
        blocos_cred.append('F100-CONTRI')

    # Blocos onde achou mas sem crédito (CST fora 50-67)
    blocos_nc = []
    if a100_no_credit:
        blocos_nc.append('A100-CONTRI')
    if f100_no_credit:
        blocos_nc.append('F100-CONTRI')

    # Blocos onde achou só pela NF (informativo — CNPJ não confirmado)
    blocos_nf_only = []
    if val_efd is not None and efd_via_nf_only:
        blocos_nf_only.append('C100-EFD')

    def _fmt_blocos(lst):
        return lst[0] if len(lst) == 1 else ', '.join(lst[:-1]) + ' e ' + lst[-1]

    # Nada encontrado em lugar nenhum
    if not blocos_cred and not blocos_nc and not blocos_nf_only:
        return {
            'analise_text': 'Não localizado nos blocos de crédito',
            'analise_style': 'analise_err',
            'soma_partidas_text': '',
            'diferenca_value': None,
        }

    # Achou só por NF (CNPJ não confirmado) — informativo, evita duplicidade
    if not blocos_cred and not blocos_nc and blocos_nf_only:
        return {
            'analise_text': f'Encontrado em {_fmt_blocos(blocos_nf_only)} (NF localizada, CNPJ não confirmado)',
            'analise_style': 'analise_info',
            'soma_partidas_text': '',
            'diferenca_value': None,
        }

    # Achou só sem crédito (nada com crédito)
    if not blocos_cred and blocos_nc:
        sufixo_nf = f' · também em {_fmt_blocos(blocos_nf_only)} (NF localizada)' if blocos_nf_only else ''
        return {
            'analise_text': f'Encontrado em {_fmt_blocos(blocos_nc)} sem crédito{sufixo_nf}',
            'analise_style': 'analise_info',
            'soma_partidas_text': '',
            'diferenca_value': None,
        }

    # Achou com crédito — somar só os valores que TÊM crédito (ignora os no_credit)
    soma_sped = 0.0
    if val_efd is not None:
        soma_sped += val_efd
    if val_a100 is not None and not a100_no_credit:
        soma_sped += val_a100
    if val_f100 is not None and not f100_no_credit:
        soma_sped += val_f100
    diff = vlr_partida - soma_sped

    blocos_txt = _fmt_blocos(blocos_cred)
    # Sufixo opcional pra blocos sem crédito e/ou NF-only quando também há crédito
    sufixos = []
    if blocos_nc:
        sufixos.append(f'também em {_fmt_blocos(blocos_nc)} sem crédito')
    if blocos_nf_only:
        sufixos.append(f'NF também em {_fmt_blocos(blocos_nf_only)}')
    sufixo_nc = f' ({"; ".join(sufixos)})' if sufixos else ''

    # 1) Match individual (Vlr Partida ≈ Σ SPED com crédito)
    if abs(diff) <= TOLERANCIA_VALOR:
        return {
            'analise_text': f'Encontrado em {blocos_txt}{sufixo_nc}',
            'analise_style': 'analise_ok',
            'soma_partidas_text': '',
            'diferenca_value': None,
        }

    # 2) Match por soma das partidas da Razão
    if (partidas_match_efd or partidas_match_a100) and partidas_count > 1 and total_partidas is not None:
        return {
            'analise_text': f'Encontrado em {blocos_txt}{sufixo_nc}',
            'analise_style': 'analise_ok',
            'soma_partidas_text': f'Fecha somando as {partidas_count} partidas ({_fmt_brl(total_partidas)})',
            'diferenca_value': None,
        }

    # 3) Divergência real
    return {
        'analise_text': f'Encontrado em {blocos_txt} com diferença{sufixo_nc}',
        'analise_style': 'analise_warn',
        'soma_partidas_text': '',
        'diferenca_value': round(diff, 2),
    }


# ============================================================
# CONSTRUÇÃO DO XLSX DE SAÍDA (com padrão visual EFCT)
# ============================================================
COLS_ORIGINAIS = 15  # Cols 1-15 da Razão original (até a NF)
NOVAS_COLS = [
    'Número da Nota',           # 16
    'Razão Social Extraída',    # 17 — do Complemento Histórico (via tokenização)
    'CNPJ',                     # 18 — identificado via cascata (4 camadas)
    'Razão Social Oficial',     # 19 — da BrasilAPI/CNPJa (consulta do CNPJ)
    '1º CNAE',                  # 20
    '2º CNAE',                  # 21
    'C100 - EFD FISCAL',        # 22
    'A100 - CONTRI',            # 23
    'F100 - CONTRI',            # 24
    'Fecha somando partidas',   # 25 — preenchida só quando match é por soma agregada
    'Valor da diferença',       # 26 — preenchida só quando há divergência (com sinal)
    'ANÁLISE DO CRUZAMENTO',    # 27 — status enxuto pra filtrar
]
TOTAL_COLS = COLS_ORIGINAIS + len(NOVAS_COLS)  # 27

COL_WIDTHS = [
    16, 9, 9, 11, 13, 9, 10, 38, 10, 13, 5, 9, 9, 38, 10,         # originais 1-15
    14, 38, 20, 38, 48, 48, 16, 16, 16, 44, 18, 38                 # novas 16-27 (12 cols)
]


class _StreamSheet:
    """Faz uma planilha openpyxl em modo write_only (grava em disco em streaming,
    RAM quase constante) se comportar como uma planilha normal indexada por célula.

    Bufferiza a linha atual e dá flush (append) quando muda de linha — assim o
    código de escrita continua usando `ws.cell(row, column, value)` + apply_style
    igual ao modo normal, mas o pico de memória não cresce com o nº de linhas.
    Atributos de planilha (freeze_panes, auto_filter, column_dimensions, etc.) são
    repassados pra planilha real.
    """
    _PROXY = {'freeze_panes', 'auto_filter', 'sheet_view',
              'column_dimensions', 'row_dimensions', 'title'}

    def __init__(self, ws):
        object.__setattr__(self, 'ws', ws)
        object.__setattr__(self, '_row', None)
        object.__setattr__(self, '_cells', {})

    def __setattr__(self, name, value):
        if name in _StreamSheet._PROXY:
            setattr(self.ws, name, value)
        else:
            object.__setattr__(self, name, value)

    def __getattr__(self, name):
        return getattr(self.ws, name)

    def cell(self, row, column, value=None):
        if self._row is None:
            self._row = row
        elif row != self._row:
            self._flush()
            self._row = row
        c = WriteOnlyCell(self.ws, value=value)
        self._cells[column] = c
        return c

    def _flush(self):
        if not self._cells:
            return
        maxc = max(self._cells)
        self.ws.append([self._cells.get(c) for c in range(1, maxc + 1)])
        object.__setattr__(self, '_cells', {})

    def done(self):
        """Dá flush da última linha bufferizada. Chamar ao terminar a planilha."""
        self._flush()


def build_workbook(df_razao, enriched_list, agg_efd, agg_a100, agg_f100, cnae_results,
                   razao_partidas_sum=None, razao_partidas_count=None,
                   agg_a100_nc=None, agg_f100_nc=None,
                   nf_idx_efd=None,
                   efd_stream=None, a100_stream=None, f100_stream=None,
                   col_efd=None, col_a100=None, col_f100=None):
    """Constrói o workbook com Razão cruzada + Pendências + 3 abas dos blocos SPED.

    `agg_a100_nc` e `agg_f100_nc` são aggregates de NFs SEM crédito (CST fora 50-67).
    Quando uma linha não acha match com crédito mas acha sem crédito, a célula é
    preenchida em estilo informativo (italic navy) e a Análise indica "sem crédito".

    `nf_idx_efd` é o índice NF-only do C100-EFD — usado pra mostrar a presença da NF
    no Fiscal mesmo quando o cruzamento estrito (NF+CNPJ) falhou. Evita que a usuária
    conte essa NF como benefício em duas análises diferentes (duplicidade).

    `df_efd_raw / df_a100_raw / df_f100_raw` são os DataFrames brutos dos SPEDs.
    Quando passados, são escritos como abas auditáveis (BLOCO_C100-EFD, BLOCO_A100,
    BLOCO_F100) e as colunas de valor cruzado passam a usar fórmulas SUMIFS apontando
    pra essas abas — assim a usuária clica na célula e vê exatamente como o valor foi
    encontrado, podendo navegar pra origem.
    """
    razao_partidas_sum = razao_partidas_sum or {}
    razao_partidas_count = razao_partidas_count or {}
    agg_a100_nc = agg_a100_nc or {}
    agg_f100_nc = agg_f100_nc or {}
    nf_idx_efd = nf_idx_efd or {}
    # write_only = gravação em streaming (RAM quase constante mesmo com blocos enormes)
    wb = Workbook(write_only=True)
    ws = _StreamSheet(wb.create_sheet('RAZÃO CONTABIL'))
    styles = make_styles()

    # Nomes das abas auxiliares (referenciadas pelas fórmulas SUMIFS)
    SHEET_EFD = 'BLOCO_C100-EFD'
    SHEET_A100 = 'BLOCO_A100'
    SHEET_F100 = 'BLOCO_F100'
    use_efd_formula = efd_stream is not None
    use_a100_formula = a100_stream is not None
    use_f100_formula = f100_stream is not None

    # ----- Ajustes da aba (write_only: DEFINIR ANTES de escrever linhas) -----
    # Larguras, altura do cabeçalho, gridlines e congelamento são gravados no
    # início do XML da planilha — então têm que vir antes do 1º append.
    for col_idx, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    # ----- Cabeçalho -----
    header_src = list(df_razao.columns)[:COLS_ORIGINAIS]
    # Garante 15 cols mesmo se o original tiver menos
    while len(header_src) < COLS_ORIGINAIS:
        header_src.append('')
    full_header = header_src + NOVAS_COLS
    for col_idx, name in enumerate(full_header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        if col_idx <= COLS_ORIGINAIS:
            apply_style(cell, styles['header'])
        else:
            apply_style(cell, styles['new_header'])

    # ----- Stats e pendências -----
    stats = {
        'linhas_razao': 0,
        'match_efd': 0, 'match_a100': 0, 'match_f100': 0,
        'match_any': 0, 'no_match': 0,
        'div_efd': 0, 'div_a100': 0, 'div_f100': 0,
        'sem_divergencia': 0,
        'amb_unresolved': 0, 'not_found': 0, 'cnpj_missing': 0,
        'amb_resolved': 0,
    }
    pendencias = {}  # cnpj → {cnpj, nome, cnae1, cnae2, total, count}

    # ----- Linhas de dados -----
    for i, (row_data, meta) in enumerate(zip(df_razao.itertuples(index=False), enriched_list), start=2):
        stats['linhas_razao'] += 1
        # Cols 1-15 originais
        for c in range(COLS_ORIGINAIS):
            val = row_data[c] if c < len(row_data) else None
            if val is not None and isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=c + 1, value=val)
            apply_style(cell, styles['default'])

        # Resolve display dos campos enriquecidos
        nf_ext = meta['nf_ext']
        st = nf_ext['status']
        if st in ('CONFIDENT', 'CONFIDENT_ALT', 'AMBIGUOUS_RESOLVED', 'FROM_COL15'):
            nf_display = nf_ext['value']
            nf_style = 'info'
            nf_value = nf_ext['value']
        elif st == 'AMBIGUOUS':
            nf_display = '⚠ AMBÍGUO: ' + '/'.join(nf_ext['candidates'])
            nf_style = 'alert'
            nf_value = None
            stats['amb_unresolved'] += 1
        elif st == 'NOT_FOUND':
            nf_display = '⚠ NÃO IDENTIFICADO'
            nf_style = 'alert'
            nf_value = None
            stats['not_found'] += 1
        else:
            nf_display = ''
            nf_style = 'info'
            nf_value = None
        if st == 'AMBIGUOUS_RESOLVED':
            stats['amb_resolved'] += 1
        if st == 'FROM_COL15':
            stats['from_col15'] = stats.get('from_col15', 0) + 1

        cnpj = meta['cnpj']
        cnpj_via = meta.get('cnpj_via') or ''
        if cnpj_via.startswith('NF_SPED'):
            stats['cnpj_via_nf_sped'] = stats.get('cnpj_via_nf_sped', 0) + 1
        elif cnpj_via == 'HISTORICO_SCAN':
            stats['cnpj_via_scan'] = stats.get('cnpj_via_scan', 0) + 1
        elif cnpj_via == 'DIRETO_HISTORICO':
            stats['cnpj_via_direto'] = stats.get('cnpj_via_direto', 0) + 1
        cnae_data = cnae_results.get(cnpj, {}) if cnpj else {}
        # Razão Social Extraída — vinda do Complemento Histórico (tokenização)
        razao_extraida = meta['supplier_name'] or ''
        # Razão Social Oficial — vinda da BrasilAPI/CNPJa (consulta do CNPJ)
        razao_oficial = cnae_data.get('razao_social') or ''
        cnpj_display = format_cnpj(cnpj) if cnpj else '⚠ CNPJ NÃO ENCONTRADO' if meta else ''
        if not cnpj:
            stats['cnpj_missing'] += 1
        cnae1 = cnae_data.get('cnae1_desc') or ''
        cnae2 = cnae_data.get('cnae2_desc') or ''

        # Col 16: NF Extraída
        cell = ws.cell(row=i, column=16, value=nf_display)
        apply_style(cell, styles[nf_style])

        # Col 17: Razão Social Extraída (do histórico)
        cell = ws.cell(row=i, column=17, value=razao_extraida)
        apply_style(cell, styles['info'])

        # Col 18: CNPJ
        cell = ws.cell(row=i, column=18, value=cnpj_display)
        apply_style(cell, styles['info'] if cnpj else styles['alert'])

        # Col 19: Razão Social Oficial (da consulta do CNPJ)
        cell = ws.cell(row=i, column=19, value=razao_oficial)
        apply_style(cell, styles['info'])

        # Col 20: 1º CNAE (descrição completa)
        cell = ws.cell(row=i, column=20, value=cnae1)
        apply_style(cell, styles['info'])

        # Col 21: 2º CNAE
        cell = ws.cell(row=i, column=21, value=cnae2)
        apply_style(cell, styles['info'])

        # Triple Check
        vlr_partida = to_number(row_data[9]) if len(row_data) > 9 else 0  # col 10 Vlr Partida (0-indexed 9)
        periodo = to_month_year(row_data[1]) if len(row_data) > 1 else ''  # col 2 Período

        # Soma de partidas da Razão para esta NF+CNPJ (pra detectar match por soma)
        partida_key = f'{nf_value}|{cnpj}' if (nf_value and cnpj) else None
        total_partidas = razao_partidas_sum.get(partida_key, vlr_partida) if partida_key else vlr_partida
        partidas_count = razao_partidas_count.get(partida_key, 1) if partida_key else 1

        # C100-EFD: match por NF + CNPJ — soma itens da nota.
        # Se Σ itens bate com Vlr Partida → match individual (verde).
        # Se Σ itens não bate com partida mas bate com Σ TODAS as partidas dessa NF → match por soma (verde, sinalizado).
        val_efd = None
        diverge_efd = False
        partidas_match_efd = False
        efd_via_nf_only = False  # NF está no C100-EFD mas CNPJ não foi confirmado
        if nf_value and cnpj:
            hit = agg_efd.get(f'{nf_value}|{cnpj}')
            if hit:
                val_efd = round(hit['total'], 2)
                diff_ind = abs(val_efd - vlr_partida)
                diff_tot = abs(val_efd - total_partidas)
                if diff_ind <= TOLERANCIA_VALOR:
                    diverge_efd = False
                elif partidas_count > 1 and diff_tot <= TOLERANCIA_VALOR:
                    diverge_efd = False
                    partidas_match_efd = True
                else:
                    diverge_efd = True
                stats['match_efd'] += 1
                if diverge_efd:
                    stats['div_efd'] += 1
                if partidas_match_efd:
                    stats['partidas_match_efd'] = stats.get('partidas_match_efd', 0) + 1

        # FALLBACK C100-EFD (CNPJ não confirmado) — SÓ quando é SEGURO:
        # a NF tem UMA única ocorrência no C100 (um único CNPJ) E o valor dela
        # bate com a Vlr Partida (ou com a soma das partidas, se a NF foi dividida).
        # Nunca mais somo notas de fornecedores diferentes que só compartilham o
        # número da NF — era isso que trazia valores que não fechavam.
        if val_efd is None and nf_value:
            candidates = nf_idx_efd.get(nf_value, [])
            if len(candidates) == 1:
                only_val = round(candidates[0][1]['total'], 2)
                bate = (abs(only_val - vlr_partida) <= TOLERANCIA_VALOR
                        or (partidas_count > 1 and abs(only_val - total_partidas) <= TOLERANCIA_VALOR))
                if bate:
                    val_efd = only_val
                    efd_via_nf_only = True
                    stats['efd_via_nf_only'] = stats.get('efd_via_nf_only', 0) + 1

        # A100: mesma lógica de C100-EFD. Se não achar com crédito, tenta sem crédito.
        val_a100 = None
        diverge_a100 = False
        partidas_match_a100 = False
        a100_no_credit = False  # NF está no A100 mas em CST fora 50-67
        if nf_value and cnpj:
            hit = agg_a100.get(f'{nf_value}|{cnpj}')
            if hit:
                val_a100 = round(hit['total'], 2)
                diff_ind = abs(val_a100 - vlr_partida)
                diff_tot = abs(val_a100 - total_partidas)
                if diff_ind <= TOLERANCIA_VALOR:
                    diverge_a100 = False
                elif partidas_count > 1 and diff_tot <= TOLERANCIA_VALOR:
                    diverge_a100 = False
                    partidas_match_a100 = True
                else:
                    diverge_a100 = True
                stats['match_a100'] += 1
                if diverge_a100:
                    stats['div_a100'] += 1
                if partidas_match_a100:
                    stats['partidas_match_a100'] = stats.get('partidas_match_a100', 0) + 1
            else:
                # Fallback: NF presente no A100 sem crédito (CST fora 50-67)
                hit_nc = agg_a100_nc.get(f'{nf_value}|{cnpj}')
                if hit_nc:
                    val_a100 = round(hit_nc['total'], 2)
                    a100_no_credit = True
                    stats['no_credit_a100'] = stats.get('no_credit_a100', 0) + 1

        val_f100 = None
        diverge_f100 = False
        f100_no_credit = False
        if cnpj and periodo:
            m = find_f100_match(agg_f100, cnpj, periodo, vlr_partida)
            if m and m['exact_match']:
                val_f100 = round(m['total'], 2)
                diverge_f100 = abs(val_f100 - vlr_partida) > TOLERANCIA_VALOR
                stats['match_f100'] += 1
                if diverge_f100:
                    stats['div_f100'] += 1
            else:
                # Fallback: F100 sem crédito (CST fora 50-67)
                m_nc = find_f100_match(agg_f100_nc, cnpj, periodo, vlr_partida)
                if m_nc and m_nc['exact_match']:
                    val_f100 = round(m_nc['total'], 2)
                    f100_no_credit = True
                    stats['no_credit_f100'] = stats.get('no_credit_f100', 0) + 1

        # ── Colunas dos blocos: TODA célula com valor vira FÓRMULA auditável ──
        # Em cada aba BLOCO_*: col A = Chave (CNPJ|NF ou CNPJ|Período), col B = Valor
        # somado, col C = Com Crédito (1/0), col D (só F100) = Vlr Operação. A fórmula
        # casa por essas colunas fixas; ROUND(...;2) pra bater com o valor calculado.
        chave_nf = f'R{i}&"|"&P{i}'  # CNPJ (col R) | NF (col P) — chave de texto

        # Col 22: C100 - EFD FISCAL
        cell = ws.cell(row=i, column=22, value=val_efd)
        if val_efd is not None and use_efd_formula:
            if not efd_via_nf_only:
                # Match estrito: chave CNPJ|NF (col A) → soma Valor (col B)
                cell.value = (
                    f"=IFERROR(ROUND(SUMIFS('{SHEET_EFD}'!B:B,"
                    f"'{SHEET_EFD}'!A:A,{chave_nf}),2),\"\")"
                )
            else:
                # Fallback NF-only: soma todas as ocorrências dessa NF (col C), qualquer CNPJ
                cell.value = (
                    f"=IFERROR(ROUND(SUMIF('{SHEET_EFD}'!C:C,P{i},"
                    f"'{SHEET_EFD}'!B:B),2),\"\")"
                )
        if efd_via_nf_only:
            apply_style(cell, styles['value_info'])  # NF localizada sem confirmação de CNPJ
        else:
            apply_style(cell, styles['value_warn'] if diverge_efd else styles['value_ok'])

        # Col 23: A100 - CONTRI — chave CNPJ|NF (A) + Valor (B) + Com Crédito (C)
        cell = ws.cell(row=i, column=23, value=val_a100)
        if val_a100 is not None and use_a100_formula:
            cred_flag = 0 if a100_no_credit else 1
            cell.value = (
                f"=IFERROR(ROUND(SUMIFS('{SHEET_A100}'!B:B,"
                f"'{SHEET_A100}'!A:A,{chave_nf},"
                f"'{SHEET_A100}'!C:C,{cred_flag}),2),\"\")"
            )
        if a100_no_credit:
            apply_style(cell, styles['value_info'])
        else:
            apply_style(cell, styles['value_warn'] if diverge_a100 else styles['value_ok'])

        # Col 24: F100 - CONTRI — chave CNPJ|Período (A) + Valor (B) + Vlr Op tolerância (D) + Crédito (C)
        cell = ws.cell(row=i, column=24, value=val_f100)
        if val_f100 is not None and use_f100_formula and periodo:
            cred_flag = 0 if f100_no_credit else 1
            lo = vlr_partida - TOLERANCIA_VALOR
            hi = vlr_partida + TOLERANCIA_VALOR
            chave_per = f'R{i}&"|{periodo}"'  # CNPJ (col R) | Período (literal)
            cell.value = (
                f"=IFERROR(ROUND(SUMIFS('{SHEET_F100}'!B:B,"
                f"'{SHEET_F100}'!A:A,{chave_per},"
                f"'{SHEET_F100}'!D:D,\">={lo:.2f}\","
                f"'{SHEET_F100}'!D:D,\"<={hi:.2f}\","
                f"'{SHEET_F100}'!C:C,{cred_flag}),2),\"\")"
            )
        if f100_no_credit:
            apply_style(cell, styles['value_info'])
        else:
            apply_style(cell, styles['value_warn'] if diverge_f100 else styles['value_ok'])

        analise = build_analise(
            vlr_partida, val_efd, val_a100, val_f100,
            partidas_match_efd=partidas_match_efd,
            partidas_match_a100=partidas_match_a100,
            total_partidas=total_partidas,
            partidas_count=partidas_count,
            a100_no_credit=a100_no_credit,
            f100_no_credit=f100_no_credit,
            efd_via_nf_only=efd_via_nf_only,
        )

        # Col 25: Fecha somando partidas (só preenchida quando match foi via soma)
        cell = ws.cell(row=i, column=25, value=analise['soma_partidas_text'] or None)
        apply_style(cell, styles['info'])

        # Col 26: Valor da diferença (só preenchida quando há divergência)
        diff_val = analise['diferenca_value']
        cell = ws.cell(row=i, column=26, value=diff_val if diff_val is not None else None)
        if diff_val is not None:
            apply_style(cell, styles['value_warn'])
            cell.number_format = 'R$ #,##0.00;-R$ #,##0.00'
        else:
            apply_style(cell, styles['default'])

        # Col 27: ANÁLISE DO CRUZAMENTO (status enxuto pra filtrar)
        cell = ws.cell(row=i, column=27, value=analise['analise_text'])
        apply_style(cell, styles[analise['analise_style']])

        # Stats e pendências
        if analise['analise_style'] == 'analise_ok':
            stats['sem_divergencia'] += 1
        if val_efd is not None or val_a100 is not None or val_f100 is not None:
            stats['match_any'] += 1
        else:
            stats['no_match'] += 1
            # Pendência: agrega por CNPJ (ou marker '__NO_CNPJ__')
            # Prefere a Razão Social Oficial (BrasilAPI) — mais confiável que a extraída
            nome_para_pendencia = razao_oficial or razao_extraida
            pkey = cnpj or '__NO_CNPJ__'
            p = pendencias.get(pkey)
            if p is None:
                p = {
                    'cnpj': cnpj or '',
                    'nome': nome_para_pendencia,
                    'cnae1': cnae1,
                    'cnae2': cnae2,
                    'total': 0.0,
                    'count': 0,
                }
                pendencias[pkey] = p
            p['total'] += vlr_partida
            p['count'] += 1
            if not p['nome'] and nome_para_pendencia:
                p['nome'] = nome_para_pendencia

    # ----- Ajustes finais da aba RAZÃO (só o autofiltro depende do total) -----
    ws.auto_filter.ref = f'A1:{get_column_letter(TOTAL_COLS)}{stats["linhas_razao"] + 1}'
    ws.done()  # flush da última linha bufferizada

    # ----- Aba PENDÊNCIAS DE CRUZAMENTO -----
    pend_total_geral = 0.0
    pend_list = sorted(pendencias.values(), key=lambda x: x['total'], reverse=True)
    if pend_list:
        ws2 = _StreamSheet(wb.create_sheet('PENDÊNCIAS DE CRUZAMENTO'))
        # Ajustes da aba ANTES de escrever (write_only)
        for col_idx, w in enumerate([22, 42, 48, 48, 24, 16], start=1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = w
        ws2.row_dimensions[1].height = 32
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = 'A2'
        pend_header = [
            'CNPJ', 'Razão Social', '1º CNAE', '2º CNAE',
            'Valor Total Não Encontrado', 'Qtd Lançamentos',
        ]
        for col_idx, name in enumerate(pend_header, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=name)
            apply_style(cell, styles['new_header'])

        for i, p in enumerate(pend_list, start=2):
            ws2.cell(row=i, column=1, value=format_cnpj(p['cnpj']) if p['cnpj'] else '⚠ Sem CNPJ')
            ws2.cell(row=i, column=2, value=p['nome'] or '(sem identificação)')
            ws2.cell(row=i, column=3, value=p['cnae1'] or '')
            ws2.cell(row=i, column=4, value=p['cnae2'] or '')
            ws2.cell(row=i, column=5, value=round(p['total'], 2))
            ws2.cell(row=i, column=6, value=p['count'])
            pend_total_geral += p['total']

            apply_style(ws2.cell(row=i, column=1), styles['info'] if p['cnpj'] else styles['alert'])
            apply_style(ws2.cell(row=i, column=2), styles['info'])
            apply_style(ws2.cell(row=i, column=3), styles['info'])
            apply_style(ws2.cell(row=i, column=4), styles['info'])
            apply_style(ws2.cell(row=i, column=5), styles['value_warn'])
            apply_style(ws2.cell(row=i, column=6), styles['info'])

        # Linha de total
        total_row = len(pend_list) + 2
        ws2.cell(row=total_row, column=4, value='TOTAL GERAL →')
        ws2.cell(row=total_row, column=5, value=round(pend_total_geral, 2))
        ws2.cell(row=total_row, column=6, value=sum(p['count'] for p in pend_list))
        for c in [4, 5, 6]:
            apply_style(ws2.cell(row=total_row, column=c), styles['new_header'])
        ws2.cell(row=total_row, column=5).number_format = '#,##0.00'

        ws2.auto_filter.ref = f'A1:F{len(pend_list) + 1}'
        ws2.done()  # flush da última linha (total geral)

    stats['pendencias'] = len(pend_list)
    stats['pendencias_total'] = round(pend_total_geral, 2)

    # ----- Abas dos blocos SPED (COMPLETAS, lidas em STREAMING) -----
    # Cada aba traz as colunas-chave fixas no início (usadas pelos SUMIFS) seguidas
    # de TODAS as colunas originais do arquivo, lidas direto do arquivo em streaming
    # (sem segurar tudo na RAM) — cabe no plano free mesmo com arquivos gigantes.
    if efd_stream is not None:
        _write_bloco_full(wb, SHEET_EFD, efd_stream, col_efd or {}, styles, kind='efd')
    if a100_stream is not None:
        _write_bloco_full(wb, SHEET_A100, a100_stream, col_a100 or {}, styles, kind='a100')
    if f100_stream is not None:
        _write_bloco_full(wb, SHEET_F100, f100_stream, col_f100 or {}, styles, kind='f100')

    return wb, stats


def _cst_to_int(cst_raw):
    """Converte CST pra int (ou None). Aceita '50', '50.0', 50 etc."""
    if cst_raw is None or (isinstance(cst_raw, float) and pd.isna(cst_raw)):
        return None
    try:
        return int(float(str(cst_raw).strip()))
    except (ValueError, TypeError):
        return None


def _clean_cell(v):
    """Prepara um valor cru do pandas pra escrever no openpyxl (NaN/NaT → None)."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _write_bloco_full(wb, sheet_name, file_stream, col_map, styles, kind):
    """Escreve a aba de bloco COMPLETA lendo o arquivo em STREAMING (openpyxl
    read_only) — nunca segura o bloco inteiro na RAM. Funciona com arquivos enormes
    (dezenas de milhares de linhas × dezenas de colunas) no plano free.

    Colunas-chave fixas no início (usadas pelas FÓRMULAS) + TODAS as colunas
    originais do arquivo importado (pra consulta/filtro):
      efd:   A=Chave(CNPJ|NF) · B=Valor(num) · C=NF(norm)
      a100:  A=Chave(CNPJ|NF) · B=Valor(num) · C=Com Crédito(1/0)
      f100:  A=Chave(CNPJ|Período) · B=Valor(num) · C=Com Crédito(1/0) · D=Vlr Operação
    """
    from openpyxl import load_workbook as _load_ro
    ws = wb.create_sheet(sheet_name)

    if kind == 'efd':
        key_headers = ['Chave (CNPJ|NF)', 'Valor (p/ fórmula)', 'NF (norm.)']
    elif kind == 'a100':
        key_headers = ['Chave (CNPJ|NF)', 'Valor (p/ fórmula)', 'Com Crédito (1=sim)']
    else:  # f100
        key_headers = ['Chave (CNPJ|Período)', 'Valor (p/ fórmula)',
                       'Com Crédito (1=sim)', 'Vlr Operação (norm.)']
    n_key = len(key_headers)
    sep_col = n_key + 1
    orig_start = n_key + 2
    val_col = 2
    vlrop_col = 4 if kind == 'f100' else None

    i_cnpj = col_map.get('cnpj'); i_nf = col_map.get('nf'); i_val = col_map.get('valor')
    i_cst = col_map.get('cst'); i_per = col_map.get('periodo'); i_op = col_map.get('vlr_operacao')

    def g(vals, idx):
        return vals[idx] if (idx is not None and idx < len(vals)) else None

    # Abre o arquivo em modo read_only (streaming, lazy) — não carrega tudo na RAM
    if hasattr(file_stream, 'seek'):
        file_stream.seek(0)
    wb_in = _load_ro(file_stream, read_only=True, data_only=True)
    src = wb_in.active
    rows_iter = src.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        header_row = ()
    orig_headers = [(str(h) if h is not None else '') for h in header_row]
    full_header = key_headers + ['↓ BLOCO ORIGINAL ↓'] + orig_headers

    # Ajustes da aba ANTES de escrever linhas (write_only grava isso no início do XML)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions[get_column_letter(sep_col)].width = 4
    for j in range(len(orig_headers)):
        ws.column_dimensions[get_column_letter(orig_start + j)].width = 18
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = get_column_letter(orig_start) + '2'

    hdr_cells = []
    for name in full_header:
        hc = WriteOnlyCell(ws, value=name)
        apply_style(hc, styles['new_header'])
        hdr_cells.append(hc)
    ws.append(hdr_cells)

    nrows = 0
    for vals in rows_iter:
        nrows += 1
        cnpj_n = normalize_cnpj(g(vals, i_cnpj))
        cnpj_fmt = format_cnpj(cnpj_n) if len(cnpj_n) == 14 else ''
        valor = to_number(g(vals, i_val))

        if kind == 'efd':
            nf_n = normalize_nf(g(vals, i_nf))
            key_vals = [f'{cnpj_fmt}|{nf_n}', valor, nf_n]
        elif kind == 'a100':
            nf_n = normalize_nf(g(vals, i_nf))
            cst_i = _cst_to_int(g(vals, i_cst))
            cred = 1 if (cst_i is not None and 50 <= cst_i <= 67) else 0
            key_vals = [f'{cnpj_fmt}|{nf_n}', valor, cred]
        else:  # f100
            per = to_month_year(g(vals, i_per))
            cst_i = _cst_to_int(g(vals, i_cst))
            cred = 1 if (cst_i is not None and 50 <= cst_i <= 67) else 0
            vlr_op = to_number(g(vals, i_op))
            key_vals = [f'{cnpj_fmt}|{per}', valor, cred, vlr_op]

        row_out = key_vals + [None] + [_clean_cell(v) for v in vals]
        cB = WriteOnlyCell(ws, value=row_out[val_col - 1]); cB.number_format = '#,##0.00'
        row_out[val_col - 1] = cB
        if vlrop_col:
            cD = WriteOnlyCell(ws, value=row_out[vlrop_col - 1]); cD.number_format = '#,##0.00'
            row_out[vlrop_col - 1] = cD
        ws.append(row_out)

    wb_in.close()
    ws.auto_filter.ref = f'A1:{get_column_letter(len(full_header))}{max(nrows + 1, 1)}'
    log.info('Aba "%s" (streaming): %s linhas × %s colunas', sheet_name, nrows, len(full_header))


def _ordena_periodos(pers):
    """Ordena períodos MM/AAAA cronologicamente."""
    def _k(p):
        try:
            mm, yy = p.split('/')
            return (int(yy), int(mm))
        except Exception:
            return (0, 0)
    return sorted(pers, key=_k)


def _faixa_periodos(pers):
    """'MM/AAAA a MM/AAAA' (ou só um, ou vazio)."""
    if not pers:
        return ''
    o = _ordena_periodos(pers)
    return o[0] if len(o) == 1 else f'{o[0]} a {o[-1]}'


# ============================================================
# FUNÇÃO PRINCIPAL
# ============================================================
def preview_padrao(razao_stream, estrategia=0, exemplo_historico=None, exemplo_razao=None):
    """Lê só a Razão e devolve uma amostra da extração (NF + Razão Social) pra a
    usuária CONFERIR o padrão antes de processar tudo. Rápido — só lê a Razão."""
    df_razao = _safe_read_excel(razao_stream, friendly_name='Razão Contábil',
                                header=0, dtype=object)
    return preview_extracao(df_razao, strategy=int(estrategia or 0),
                            exemplo_historico=exemplo_historico, exemplo_razao=exemplo_razao)


def processar_cruzamento(
    razao_stream, c100efd_stream, a100_stream=None, f100_stream=None, estrategia=0,
    exemplo_historico=None, exemplo_razao=None
):
    """
    Entrypoint chamado pelo Flask.
    `a100_stream` e `f100_stream` são OPCIONAIS — se a empresa não tem esses blocos,
    o processamento segue só com Razão + C100-EFD (aggregates A100/F100 ficam vazios).
    A consulta CNPJ → Razão Social Oficial + CNAEs via BrasilAPI/CNPJa é SEMPRE feita
    quando há CNPJs identificados (não há opção de pular).
    Retorna (bytes_xlsx, dict_stats).
    """
    log.info('Lendo Razão Contábil…')
    df_razao = _safe_read_excel(razao_stream, friendly_name='Razão Contábil',
                                header=0, dtype=object)
    log.info('Razão: %s linhas lidas', len(df_razao))

    # ─────────────────────────────────────────────────────────────────────
    # FILTRO: descarta lançamentos de "Encerramento" se houver coluna Tipo
    # ─────────────────────────────────────────────────────────────────────
    # Procura coluna "Tipo de Lançamento" (com variações: Tipo Lanc, Tipo, etc)
    tipo_col = None
    for col in df_razao.columns:
        col_norm = _normalize_for_keyword(str(col))
        if 'TIPO' in col_norm and 'LANC' in col_norm:
            tipo_col = col
            break

    if tipo_col:
        def _eh_encerramento(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return False
            s = _normalize_for_keyword(str(v))
            return 'ENCERR' in s
        mask_encerr = df_razao[tipo_col].apply(_eh_encerramento)
        removidas = int(mask_encerr.sum())
        if removidas > 0:
            df_razao = df_razao[~mask_encerr].reset_index(drop=True)
            log.info('Filtro Tipo de Lançamento: %s linhas de ENCERRAMENTO descartadas. Restam %s.',
                     removidas, len(df_razao))
        else:
            log.info('Coluna "%s" detectada — nenhuma linha de encerramento encontrada.', tipo_col)
    else:
        log.info('Coluna "Tipo de Lançamento" não encontrada na Razão — sem filtro de encerramento.')

    # CNPJ raiz da Razão (col 1, normalmente o mesmo em todas as linhas)
    razao_company_cnpj = _most_common_cnpj(df_razao.iloc[:, 0].tolist())

    cnpj_map: dict[str, str] = {}

    # Lê os bytes de cada SPED UMA vez. Cada leitura (agregação e, depois, escrita da
    # aba do bloco) recebe uma cópia NOVA (BytesIO) — nunca reaproveita o mesmo stream,
    # senão a 2ª leitura via openpyxl vinha vazia (bug das abas de bloco em branco).
    def _bytes_de(stream):
        if stream is None:
            return None
        try:
            stream.seek(0)
        except Exception:
            pass
        return stream.read()

    efd_bytes = _bytes_de(c100efd_stream)
    a100_bytes = _bytes_de(a100_stream)
    f100_bytes = _bytes_de(f100_stream)

    log.info('Processando C100-EFD FISCAL…')
    r_efd = process_sped_block(io.BytesIO(efd_bytes), SPED_SCHEMA['c100efd'], cnpj_map,
                               friendly_name='C100-EFD FISCAL')
    agg_efd_raw = r_efd['agg']
    efd_company_cnpj = r_efd['company_cnpj']
    col_efd = r_efd['col_map']
    periodos_efd = r_efd['periodos']
    log.info('C100-EFD: %s linhas · %s chaves agregadas', r_efd['total'], len(agg_efd_raw))

    if a100_bytes is not None:
        log.info('Processando A100-CONTRI…')
        r_a100 = process_sped_block(io.BytesIO(a100_bytes), SPED_SCHEMA['a100'], cnpj_map,
                                    friendly_name='A100-CONTRI')
        agg_a100_raw = r_a100['agg']
        agg_a100_nc = r_a100['agg_no_cred'] or {}
        a100_company_cnpj = r_a100['company_cnpj']
        col_a100 = r_a100['col_map']
        periodos_a100 = r_a100['periodos']
        log.info('A100: %s linhas · %s sem crédito (CST fora 50-67) · %s chaves com crédito · %s chaves sem crédito',
                 r_a100['total'], r_a100['skipped'], len(agg_a100_raw), len(agg_a100_nc))
    else:
        log.info('A100-CONTRI: arquivo não fornecido — pulando bloco.')
        agg_a100_raw = {}
        agg_a100_nc = {}
        a100_company_cnpj = None
        col_a100 = {}
        periodos_a100 = set()

    if f100_bytes is not None:
        log.info('Processando F100-CONTRI…')
        r_f100 = process_sped_block(io.BytesIO(f100_bytes), SPED_SCHEMA['f100'], cnpj_map,
                                    friendly_name='F100-CONTRI')
        agg_f100_raw = r_f100['agg']
        agg_f100_nc = r_f100['agg_no_cred'] or {}
        f100_company_cnpj = r_f100['company_cnpj']
        col_f100 = r_f100['col_map']
        periodos_f100 = r_f100['periodos']
        log.info('F100: %s linhas · %s sem crédito · %s chaves com crédito · %s chaves sem crédito',
                 r_f100['total'], r_f100['skipped'], len(agg_f100_raw), len(agg_f100_nc))
    else:
        log.info('F100-CONTRI: arquivo não fornecido — pulando bloco.')
        agg_f100_raw = {}
        agg_f100_nc = {}
        f100_company_cnpj = None
        col_f100 = {}
        periodos_f100 = set()

    log.info('Mapa CNPJ: %s fornecedores', len(cnpj_map))
    # Avisos proativos quando o mapa pode estar "pobre"
    if a100_stream is None:
        log.warning('A100 não foi enviado — fornecedores de serviços (advogados, '
                    'consultorias, prestadores) provavelmente não estarão no mapa CNPJ. '
                    'Esses casos podem aparecer como "CNPJ NÃO ENCONTRADO".')
    if f100_stream is None:
        log.warning('F100 não foi enviado — operações como depreciação, frete, '
                    'receitas financeiras não serão validadas.')

    # ──────────────────────────────────────────────────────────────────────
    # VALIDAÇÃO: a RAIZ do CNPJ (8 primeiros dígitos) precisa bater entre os
    # arquivos. Filiais da MESMA empresa têm a mesma raiz e mudam só o sufixo
    # de estabelecimento (/0001 matriz, /0005 filial, etc.) — isso é PERMITIDO,
    # é a mesma empresa. Só bloqueia se a RAIZ divergir (outra empresa de fato).
    # ──────────────────────────────────────────────────────────────────────
    cnpjs_por_arquivo = {
        'Razão Contábil': razao_company_cnpj,
        'C100-EFD FISCAL': efd_company_cnpj,
        'A100-CONTRI': a100_company_cnpj,
        'F100-CONTRI': f100_company_cnpj,
    }
    log.info('CNPJs da empresa-raiz por arquivo: %s', cnpjs_por_arquivo)
    cnpjs_validos = {k: v for k, v in cnpjs_por_arquivo.items() if v}
    raizes = {k: v[:8] for k, v in cnpjs_validos.items()}  # raiz = 8 primeiros dígitos
    raizes_distintas = set(raizes.values())
    if len(raizes_distintas) > 1:
        # Raízes diferentes = empresas DIFERENTES de verdade → bloqueia
        majoritaria = max(raizes_distintas, key=lambda r: list(raizes.values()).count(r))
        divergentes = [
            f'{nome}={format_cnpj(cnpjs_validos[nome])}'
            for nome, r in raizes.items() if r != majoritaria
        ]
        exemplo = next(c for n, c in cnpjs_validos.items() if raizes[n] == majoritaria)
        msg = (
            f'⚠ Os arquivos parecem ser de EMPRESAS DIFERENTES (raiz do CNPJ diverge). '
            f'Maioria: {format_cnpj(exemplo)} (raiz {majoritaria}). '
            f'Divergente(s): {" · ".join(divergentes)}. '
            f'Confira se você não importou um SPED de outra empresa por engano. '
            f'(Filiais da mesma empresa — mesma raiz, sufixo /0001, /0002… diferente — '
            f'são aceitas normalmente.)'
        )
        raise ValueError(msg)
    elif len(raizes_distintas) == 0:
        log.warning('Nenhum CNPJ da empresa-raiz pôde ser identificado em nenhum arquivo.')
    else:
        estabelecimentos = set(cnpjs_validos.values())
        if len(estabelecimentos) > 1:
            # Mesma raiz, sufixos diferentes = matriz + filiais da MESMA empresa → ok
            log.info('✓ Mesma empresa (raiz %s) com %s estabelecimentos (matriz/filiais): %s — '
                     'cruzamento liberado.',
                     next(iter(raizes_distintas)), len(estabelecimentos),
                     ', '.join(format_cnpj(c) for c in sorted(estabelecimentos)))
        else:
            log.info('✓ CNPJ raiz validado em todos os arquivos: %s',
                     format_cnpj(next(iter(estabelecimentos))))

    # Índices NF-only pros blocos com NF — usados como último fallback pra inferir CNPJ
    # quando a Razão não fornece pistas suficientes (sem CNPJ direto, sem nome no histórico, etc).
    # Ordem de prioridade: C100-EFD (mais reliable) → A100 com crédito → A100 sem crédito.
    nf_idx_list = [
        build_nf_only_index(agg_efd_raw),
        build_nf_only_index(agg_a100_raw),
        build_nf_only_index(agg_a100_nc),
    ]
    log.info('Índices NF-only: C100-EFD %s NFs · A100 com cr %s · A100 s/ cr %s',
             len(nf_idx_list[0]), len(nf_idx_list[1]), len(nf_idx_list[2]))

    # Pré-processa Razão (extração de NF, lookup CNPJ) + agrega partidas por NF+CNPJ
    # Monta UMA vez o extrator (leitura escolhida OU padrão aprendido do exemplo) e
    # reaplica em todas as linhas.
    extrator, info_extrator = montar_extrator(estrategia, exemplo_historico, exemplo_razao)
    log.info('Extraindo NF e CNPJ da Razão… (leitura: %s)', info_extrator['label'])
    enriched_list = []
    unique_cnpjs = set()
    razao_partidas_sum = {}    # chave NF|CNPJ → soma das Vlr Partida da Razão
    razao_partidas_count = {}  # chave NF|CNPJ → quantas partidas da Razão dão nesse documento
    # Convert NaN to None na Razão pra evitar problemas no openpyxl
    df_razao = df_razao.where(pd.notna(df_razao), None)
    for row in df_razao.itertuples(index=False):
        # row[13] = Complemento Histórico (col 14, 0-indexed 13)
        # row[14] = NF original (col 15, 0-indexed 14)
        complemento = row[13] if len(row) > 13 else None
        col15 = row[14] if len(row) > 14 else None
        vlr_partida_row = to_number(row[9]) if len(row) > 9 else 0  # col 10 Vlr Partida

        # Extração (NF + Razão Social) pela leitura escolhida/aprendida
        _ex = extrator(complemento, col15)
        nf_ext = {'value': _ex['value'], 'status': _ex['status'], 'candidates': _ex['candidates']}
        supplier_name = _ex['name']

        # CNPJ — cascata GUIADA PELO NOME (estritamente consistente com o histórico):
        #   1. CNPJ direto no histórico (formatado ou 14 dígitos) — autoridade máxima
        #   2. Nome extraído → busca o CNPJ no mapa dos BLOCOS SPED (exato→prefixo→fuzzy)
        #   3. Reverse lookup: nome de fornecedor SPED aparece como texto no histórico
        #   (a busca web é a 5ª camada, mais abaixo, só pros que sobraram sem CNPJ)
        #
        # NÃO inferimos mais CNPJ pela NF nos blocos: isso trazia CNPJ de fornecedor
        # ERRADO (sem relação com o nome do histórico) e fazia a fórmula somar valores
        # incoerentes. Regra agora: CNPJ só entra se o NOME fechar com o histórico.
        lookup = None
        direct_cnpj = extract_cnpj_from_historico(complemento)
        if direct_cnpj and len(direct_cnpj) == 14:
            lookup = {'cnpj': direct_cnpj, 'via': 'DIRETO_HISTORICO'}
        else:
            lookup = lookup_cnpj(supplier_name, cnpj_map)
            if not lookup:
                lookup = lookup_cnpj_by_historico_scan(complemento, cnpj_map)
                if lookup and lookup.get('matched_name') and not supplier_name:
                    supplier_name = lookup['matched_name']

        cnpj = lookup['cnpj'] if lookup else None
        if cnpj:
            unique_cnpjs.add(cnpj)
        # Acumula partidas por NF+CNPJ (apenas se a NF foi identificada com confiança)
        nf_val = nf_ext['value'] if nf_ext['status'] in (
            'CONFIDENT', 'CONFIDENT_ALT', 'AMBIGUOUS_RESOLVED'
        ) else None
        if nf_val and cnpj:
            partida_key = f'{nf_val}|{cnpj}'
            razao_partidas_sum[partida_key] = razao_partidas_sum.get(partida_key, 0.0) + vlr_partida_row
            razao_partidas_count[partida_key] = razao_partidas_count.get(partida_key, 0) + 1
        enriched_list.append({
            'nf_ext': nf_ext,
            'supplier_name': supplier_name,
            'cnpj': cnpj,
            'cnpj_via': lookup.get('via') if lookup else None,
            'matched_name': lookup.get('matched_name') if lookup else None,
        })
    log.info('CNPJs únicos a buscar CNAE: %s', len(unique_cnpjs))

    # ─────────────────────────────────────────────────────────────────────
    # 5ª CAMADA: busca CNPJ via web (DuckDuckGo) pra fornecedores não resolvidos
    # ─────────────────────────────────────────────────────────────────────
    # Junta nomes únicos que ficaram sem CNPJ após as 4 camadas anteriores
    nomes_sem_cnpj = {}  # nome → lista de índices em enriched_list
    for idx, m in enumerate(enriched_list):
        if not m or m.get('cnpj'):
            continue
        nome = (m.get('supplier_name') or '').strip()
        if nome and len(nome) >= 8:
            nomes_sem_cnpj.setdefault(nome, []).append(idx)

    if nomes_sem_cnpj:
        log.info('5ª camada: buscando %s nomes únicos via web…', len(nomes_sem_cnpj))
        web_found = 0
        abortou = False
        for i, (nome, indices) in enumerate(nomes_sem_cnpj.items(), start=1):
            # Se TODAS as fontes já estão bloqueadas (típico em servidor/Render com
            # IP de datacenter), aborta a busca — não adianta dormir 3s por nome.
            if _all_sources_blocked():
                log.warning('5ª camada ABORTADA: todas as fontes web bloqueadas '
                            '(IP de servidor). Resolvidos %s/%s antes de abortar em %s.',
                            web_found, len(nomes_sem_cnpj), i)
                abortou = True
                break
            result = search_cnpj_by_name_web(nome)
            if result:
                cnpj_found = result['cnpj']
                for idx in indices:
                    enriched_list[idx]['cnpj'] = cnpj_found
                    enriched_list[idx]['cnpj_via'] = 'WEB_SEARCH'
                unique_cnpjs.add(cnpj_found)
                web_found += 1
            else:
                # Só faz a pausa educada se ainda há fonte viva pra tentar
                if not _all_sources_blocked():
                    time.sleep(WEB_SEARCH_THROTTLE)
            if i % 10 == 0:
                log.info('Web search progresso: %s/%s · %s achados',
                         i, len(nomes_sem_cnpj), web_found)
        if not abortou:
            log.info('5ª camada concluída: %s/%s nomes resolvidos via web',
                     web_found, len(nomes_sem_cnpj))

    # Busca CNPJ → Razão Social Oficial + CNAEs (sempre obrigatória quando há CNPJs identificados)
    cnae_results = {}
    if unique_cnpjs:
        log.info('Consultando %s CNPJs únicos na BrasilAPI/CNPJa (Razão Social Oficial + CNAEs)…',
                 len(unique_cnpjs))
        cnae_results = fetch_all_cnaes(list(unique_cnpjs))
        com_razao = sum(1 for d in cnae_results.values() if d.get('razao_social'))
        log.info('Consulta concluída: %s respostas · %s com Razão Social Oficial',
                 len(cnae_results), com_razao)

    # ─────────────────────────────────────────────────────────────────────
    # PORTÃO DE VALIDAÇÃO: CNPJ só permanece se o nome FECHAR com o histórico
    # ─────────────────────────────────────────────────────────────────────
    # Para fontes "fracas" (FUZZY/PREFIX/WEB_SEARCH), conferimos se a razão social
    # oficial (ou o nome casado no SPED) reconcilia com o nome extraído do histórico.
    # Se não reconciliar, o CNPJ é DESCARTADO — preferimos deixar em branco a trazer
    # um fornecedor errado que faria a fórmula somar valores incoerentes.
    rejeitados = 0
    for m in enriched_list:
        if not m or not m.get('cnpj'):
            continue
        via = m.get('cnpj_via') or ''
        if via in _TRUSTED_VIAS:
            continue  # fonte confiável por construção — não precisa validar
        extraido = m.get('supplier_name') or ''
        oficial = (cnae_results.get(m['cnpj'], {}) or {}).get('razao_social') or ''
        casado_sped = m.get('matched_name') or ''
        # Reconcilia contra o melhor nome disponível (oficial e/ou casado no SPED)
        nomes_ref = [n for n in (oficial, casado_sped) if n]
        if not extraido or len(normalize_name(extraido)) < 6 or not nomes_ref:
            # Não dá pra confirmar consistência de fonte fraca → descarta (seguro)
            score = 0.0
        else:
            score = max(_names_reconcile(extraido, n) for n in nomes_ref)
        if score < RECONCILE_THRESHOLD:
            log.info('CNPJ descartado (nome não fecha): histórico="%s" × oficial="%s" (via=%s, score=%.2f)',
                     extraido[:40], oficial[:40], via, score)
            m['cnpj'] = None
            m['cnpj_via'] = 'REJEITADO_NOME'
            rejeitados += 1
    if rejeitados:
        log.info('Portão de validação: %s CNPJs descartados por não fecharem com o histórico.', rejeitados)

    # Monta o workbook
    log.info('Montando workbook de saída…')
    wb, stats = build_workbook(
        df_razao, enriched_list,
        agg_efd_raw, agg_a100_raw, agg_f100_raw,
        cnae_results,
        razao_partidas_sum, razao_partidas_count,
        agg_a100_nc=agg_a100_nc, agg_f100_nc=agg_f100_nc,
        nf_idx_efd=nf_idx_list[0],
        efd_stream=io.BytesIO(efd_bytes), col_efd=col_efd,
        a100_stream=io.BytesIO(a100_bytes) if a100_bytes is not None else None, col_a100=col_a100,
        f100_stream=io.BytesIO(f100_bytes) if f100_bytes is not None else None, col_f100=col_f100,
    )
    stats['cnaes_buscados'] = len(cnae_results)

    # ─────────────────────────────────────────────────────────────────────
    # COBERTURA DE PERÍODO: Razão × blocos
    # ─────────────────────────────────────────────────────────────────────
    # Se um bloco NÃO cobre todos os meses (MM/AAAA) presentes na Razão, os
    # cruzamentos daqueles meses dariam "não localizado" indevidamente (o bloco
    # certo não foi enviado). Sinalizamos isso pro front mostrar um alerta.
    try:
        from collections import Counter
        cont = Counter()
        razao_pers = set()
        for v in df_razao.iloc[:, 1].tolist():   # col 2 = Período
            p = to_month_year(v)
            if p:
                razao_pers.add(p)
                cont[p] += 1
        if razao_pers:
            stats['periodo_razao'] = _faixa_periodos(razao_pers)
            alerta = 0
            blocos = [
                ('efd', 'C100-EFD FISCAL', periodos_efd, True),
                ('a100', 'A100-CONTRI', periodos_a100, a100_stream is not None),
                ('f100', 'F100-CONTRI', periodos_f100, f100_stream is not None),
            ]
            for nome, label, bp, enviado in blocos:
                if not enviado:
                    continue
                if not bp:
                    stats[f'cob_{nome}'] = 'sem-coluna-periodo'
                    continue
                stats[f'cob_{nome}'] = _faixa_periodos(bp)
                faltam = _ordena_periodos(razao_pers - bp)
                if faltam:
                    alerta = 1
                    linhas_fora = sum(cont[p] for p in faltam)
                    stats[f'faltam_{nome}'] = ','.join(faltam[:120])
                    stats[f'qtdfaltam_{nome}'] = len(faltam)
                    stats[f'linhasfora_{nome}'] = linhas_fora
                    log.warning('⚠ PERÍODO: %s não cobre %s mês(es) da Razão (%s lançamentos): %s',
                                label, len(faltam), linhas_fora, ', '.join(faltam))
            stats['periodo_alerta'] = alerta
            if not alerta:
                log.info('  ✓ Cobertura de período OK — todos os blocos cobrem os meses da Razão.')
    except Exception as e:
        log.warning('Não consegui calcular cobertura de período: %s', e)

    # Libera a Razão e dá um gc antes de salvar (os blocos já são streaming, não há
    # DataFrame grande na memória — o pico fica baixo mesmo em arquivos enormes).
    import gc
    df_razao = None
    gc.collect()

    # Serializa
    buf = io.BytesIO()
    wb.save(buf)
    out_bytes = buf.getvalue()
    log.info('Saída: %s bytes', len(out_bytes))

    # ─────────────────────────────────────────────────────────────────────
    # RELATÓRIO DE SANIDADE — alerta proativo se algo parece errado
    # ─────────────────────────────────────────────────────────────────────
    total_lin = stats.get('linhas_razao', 0) or 1
    cnpjs_id = total_lin - stats.get('cnpj_missing', 0)
    sem_div = stats.get('sem_divergencia', 0)
    no_match = stats.get('no_match', 0)
    pct_cnpj = round(cnpjs_id / total_lin * 100, 1)
    pct_match = round((total_lin - no_match) / total_lin * 100, 1)
    log.info('===== RELATÓRIO DE SANIDADE =====')
    log.info('  Linhas da Razão processadas: %s', total_lin)
    log.info('  CNPJs identificados:         %s (%.1f%%)', cnpjs_id, pct_cnpj)
    log.info('  Linhas com algum match SPED: %s (%.1f%%)', total_lin - no_match, pct_match)
    log.info('  Aggregates SPED criados:     EFD=%s · A100(c/cr)=%s · A100(s/cr)=%s · F100(c/cr)=%s · F100(s/cr)=%s',
             len(agg_efd_raw), len(agg_a100_raw), len(agg_a100_nc),
             len(agg_f100_raw), len(agg_f100_nc))

    # Alertas proativos
    alertas = []
    if pct_cnpj < 50:
        alertas.append(f'Apenas {pct_cnpj}% dos CNPJs identificados — pode faltar A100/F100 ou nomes muito bagunçados')
    if pct_match < 30:
        alertas.append(f'Apenas {pct_match}% das linhas com match — verifique se os SPEDs cobrem o mesmo período da Razão')
    # Detecta aggregates suspeitos (foram processadas linhas mas 0 chaves geradas)
    if a100_stream is not None and len(agg_a100_raw) == 0 and len(agg_a100_nc) == 0:
        alertas.append('A100 processado mas 0 chaves agregadas — possível incompatibilidade de layout (colunas)')
    if f100_stream is not None and len(agg_f100_raw) == 0 and len(agg_f100_nc) == 0:
        alertas.append('F100 processado mas 0 chaves agregadas — possível incompatibilidade de layout (colunas)')
    if len(agg_efd_raw) == 0:
        alertas.append('C100-EFD processado mas 0 chaves agregadas — possível incompatibilidade de layout (colunas)')
    if alertas:
        log.warning('⚠ ALERTAS DE SANIDADE:')
        for a in alertas:
            log.warning('  - %s', a)
    else:
        log.info('  ✓ Nenhum alerta de sanidade — processamento parece saudável.')

    return out_bytes, stats
